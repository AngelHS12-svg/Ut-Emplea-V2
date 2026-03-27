from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_from_directory, session, make_response
import psycopg2
from psycopg2 import pool
import os
import uuid
import random
from datetime import datetime, timedelta
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.header import Header

from app.routes.empresa_routes import empresa_bp
from app.utils.security import strip_tags, detectar_script, detectar_sqli, validar_archivo
import re

# Importación dinámica para reportes Excel si se requiere
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    pass

app = Flask(__name__, template_folder="app/Views", static_folder="public")

def validar_registro_datos(correo, password, telefono, cp=None):
    if not re.match(r"[^@]+@[^@]+\.[^@]+", correo):
        return "Correo electrónico no válido."
    
    # Validar Contraseña (mín 12, mayus, minus, num, especial)
    if len(password) < 12:
        return "La contraseña debe tener al menos 12 caracteres."
    if not re.search(r"[A-Z]", password) or not re.search(r"[a-z]", password) or \
       not re.search(r"\d", password) or not re.search(r"[@$!%*?&]", password):
        return "La contraseña debe incluir mayúscula, minúscula, número y símbolo especial."
    
    # Validar Teléfono (10 dígitos)
    if not telefono or len(telefono) != 10 or not telefono.isdigit():
        return "El teléfono debe tener exactamente 10 dígitos numéricos."
    
    # Validar CP (5 dígitos)
    if cp is not None:
        if not cp or len(cp) != 5 or not cp.isdigit():
            return "El código postal debe tener 5 dígitos."
            
    return None
app.register_blueprint(empresa_bp, url_prefix='/empresa')
app.secret_key = "secret_key_prueba"
import os
import psycopg2
from psycopg2 import pool

# ================= DATABASE CONNECTION POOL =================
class ConnectionWrapper:
    def __init__(self, conn, pool_instance=None):
        self._conn = conn
        self._pool = pool_instance

    def __getattr__(self, name):
        return getattr(self._conn, name)

    def cursor(self, *args, **kwargs):
        return self._conn.cursor(*args, **kwargs)

    def commit(self):
        return self._conn.commit()

    def rollback(self):
        return self._conn.rollback()

    def close(self):
        if self._pool:
            self._pool.putconn(self._conn)
        else:
            self._conn.close()

db_pool = None

try:
    db_pool = pool.ThreadedConnectionPool(
        1, 120,  # Aumentamos a 120 conexiones para soportar 100+ usuarios concurrentes
        host=os.getenv("DB_HOST", "localhost"),
        database=os.getenv("DB_NAME", "bolsa_trabajo_uto"),
        user=os.getenv("DB_USER", "postgres"),
        password=os.getenv("DB_PASS", "angel123"),
        port=os.getenv("DB_PORT", "5432"),
        client_encoding='utf8'
    )
    print("Pool de conexiones creado con éxito.")
except Exception as e:
    print(f"Error al crear el pool de conexiones: {e}")

def get_connection():
    if db_pool:
        # En caso de que el pool esté vacío o fallando, podríamos intentar un reconnect, 
        # pero es más fácil obtener una conexión limpia.
        try:
            conn = db_pool.getconn()
            if conn:
                return ConnectionWrapper(conn, db_pool)
        except Exception:
            pass # Si falla el pool, hacemos fallback a una conexión directa

    try:
        return ConnectionWrapper(psycopg2.connect(
            host=os.getenv("DB_HOST", "localhost"),
            database=os.getenv("DB_NAME", "bolsa_trabajo_uto"),
            user=os.getenv("DB_USER", "postgres"),
            password=os.getenv("DB_PASS", "angel123"),
            port=os.getenv("DB_PORT", "5432"),
            client_encoding='utf8'
        ))
    except UnicodeDecodeError:
        raise Exception("❌ NO SE PUDO CONECTAR A POSTGRESQL: Posiblemente la contraseña de la BD (angel123) es incorrecta para esta computadora, o el servicio PostgreSQL no se está ejecutando.")

# Configuración de uploads
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads', 'cv')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 32 * 1024 * 1024  # 32MB max

# Configuración de Correo (Gmail SMTP SSL)
app.config['MAIL_SERVER'] = 'smtp.gmail.com'
app.config['MAIL_PORT'] = 465
app.config['MAIL_USE_SSL'] = True
app.config['MAIL_USERNAME'] = 'angelhsnew123@gmail.com'
app.config['MAIL_PASSWORD'] = 'nxjwdgnrmtgdndoa' # App Password de la cuenta personal
app.config['MAIL_DEFAULT_SENDER'] = ('UT Oriental Emplea', 'angelhsnew123@gmail.com')

def enviar_email(destinatario, asunto, mensaje_texto):
    """Envía un correo electrónico utilizando smtplib."""
    try:
        msg = MIMEMultipart()
        msg['From'] = f"{Header('UT Oriental Emplea', 'utf-8')} <{app.config['MAIL_USERNAME']}>"
        msg['To'] = destinatario
        msg['Subject'] = Header(asunto, 'utf-8')
        
        # Cuerpo del mensaje
        msg.attach(MIMEText(mensaje_texto, 'plain', 'utf-8'))
        
        # Configuración del servidor con SSL
        server = smtplib.SMTP_SSL(app.config['MAIL_SERVER'], app.config['MAIL_PORT'])
        server.login(app.config['MAIL_USERNAME'], app.config['MAIL_PASSWORD'])
        server.sendmail(app.config['MAIL_USERNAME'], destinatario, msg.as_string())
        server.quit()
        print(f"📧 Correo enviado a {destinatario}: {asunto}")
        return True
    except Exception as e:
        print(f"❌ Error al enviar correo a {destinatario}: {e}")
        return False

# Configuración de Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "home"

# Seguridad: Cabeceras de protección global
@app.after_request
def add_security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    # CSP básica para permitir recursos locales y Google Fonts/CDNs conocidos
    response.headers['Content-Security-Policy'] = "default-src 'self' mailto: tel:; script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https://cdnjs.cloudflare.com; style-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https://cdnjs.cloudflare.com https://fonts.googleapis.com; font-src 'self' https://fonts.gstatic.com https://cdnjs.cloudflare.com; img-src 'self' data: https://ui-avatars.com; connect-src 'self';"
    return response

class User(UserMixin):
    def __init__(self, id_usuario, correo, id_rol, nombre_rol):
        self.id = id_usuario
        self.correo = correo
        self.id_rol = id_rol
        self.rol = nombre_rol

@login_manager.user_loader
def load_user(user_id):
    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT u.id_usuario, u.correo, u.id_rol, r.nombre 
            FROM usuarios u 
            JOIN roles r ON u.id_rol = r.id_rol 
            WHERE u.id_usuario = %s
        """, (user_id,))
        user_data = cur.fetchone()
        if user_data:
            return User(user_data[0], user_data[1], user_data[2], user_data[3])
    except Exception as e:
        print(f"Error en load_user: {e}")
    finally:
        cur.close()
        conn.close()
    return None

def release_connection(conn):
    conn.close()  # Ahora ConnectionWrapper.close() hace el putconn automáticamente


def crear_notificacion(id_usuario, tipo, mensaje, url="", asunto_p=None, cuerpo_p=None):
    conn = get_connection()
    cur = conn.cursor()
    try:
        # 1. Guardar en base de datos
        cur.execute("""
            INSERT INTO notificaciones (id_usuario, tipo, mensaje, url)
            VALUES (%s, %s, %s, %s)
        """, (id_usuario, tipo, mensaje, url))
        conn.commit()
        
        # 2. Enviar por correo electrónico
        cur.execute("SELECT correo FROM usuarios WHERE id_usuario = %s", (id_usuario,))
        user_row = cur.fetchone()
        if user_row and user_row[0]:
            destinatario = user_row[0]
            asunto = asunto_p or f"Notificación: {tipo.capitalize()}"
            link_info = f"\n\nPuedes ver más detalles aquí: http://localhost:5001{url}" if url else ""
            
            if cuerpo_p:
                cuerpo = cuerpo_p
            else:
                cuerpo = f"Hola,\n\nTienes una nueva notificación en UT Oriental Emplea:\n\n{mensaje}{link_info}\n\nSaludos,\nEl equipo de UT Oriental."
            
            enviar_email(destinatario, asunto, cuerpo)

    except Exception as e:
        if conn: conn.rollback()
        print(f"Error creando notificación: {e}")
    finally:
        cur.close()
        conn.close()

def notificar_admins(tipo, mensaje, url=""):
    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute("SELECT id_usuario FROM usuarios WHERE id_rol = 1")
        admins = cur.fetchall()
        for admin in admins:
            crear_notificacion(admin[0], tipo, mensaje, url)
    except Exception as e:
        print(f"Error notificando admins: {e}")
    finally:
        cur.close()
        conn.close()

# ================= NOTIFICACIONES =================
@app.route("/api/notificaciones", methods=["GET"])
@login_required
def obtener_notificaciones():
    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT id_notificacion, tipo, mensaje, url, fecha 
            FROM notificaciones 
            WHERE id_usuario = %s AND leida = FALSE
            ORDER BY fecha DESC
        """, (current_user.id,))
        notificaciones = [{
            "id": row[0],
            "tipo": row[1],
            "mensaje": row[2],
            "url": row[3],
            "fecha": row[4].strftime("%d/%m/%Y %H:%M") if row[4] else ""
        } for row in cur.fetchall()]
        return jsonify(notificaciones)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        cur.close()
        conn.close()

@app.route("/api/notificaciones/leer/<int:id_notificacion>", methods=["POST"])
@login_required
def leer_notificacion(id_notificacion):
    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute("""
            UPDATE notificaciones SET leida = TRUE 
            WHERE id_notificacion = %s AND id_usuario = %s
        """, (id_notificacion, current_user.id))
        conn.commit()
        return jsonify({"success": True})
    except Exception as e:
        conn.rollback()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        cur.close()
        conn.close()

# ================= HOME =================
@app.route("/")
def home():
    return render_template("home/index.html")

@app.route("/login")
def login():
    return render_template("home/index.html")

# ================= ADMIN =================
@app.route("/admin")
@login_required
def admin():
    if current_user.rol != "Administrador":
        flash("Acceso denegado: Se requiere rol de Administrador")
        return redirect(url_for("home"))
    return render_template("admi/index.html")

@app.route("/admin/manual")
@login_required
def admin_manual():
    if current_user.rol != "Administrador":
        return redirect(url_for("home"))
    return render_template("admi/manual.html")

@app.route("/admin/inicio")
@login_required
def admin_inicio():
    if current_user.rol != "Administrador":
        return redirect(url_for("home"))
    conn = get_connection()
    cur = conn.cursor()
    # Estadísticas reales
    cur.execute("SELECT COUNT(*) FROM empresas WHERE estatus = 'aprobada'")
    empresas_activas = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM candidatos")
    candidatos_total = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM vacantes WHERE estatus = 'activa'")
    vacantes_activas = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM empresas WHERE estatus = 'pendiente'")
    empresas_pendientes = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM candidatos WHERE estatus = 'pendiente'")
    candidatos_pendientes = cur.fetchone()[0]
    
    # Actividad reciente (últimas empresas y candidatos registrados)
    cur.execute("""
        SELECT e.nombre, e.fecha_registro, e.estatus FROM empresas e
        ORDER BY e.fecha_registro DESC LIMIT 5
    """)
    actividad_empresas = cur.fetchall()
    
    cur.execute("""
        SELECT c.nombre, c.fecha_registro, c.estatus FROM candidatos c
        ORDER BY c.fecha_registro DESC LIMIT 5
    """)
    actividad_candidatos = cur.fetchall()
    
    cur.close()
    conn.close()
    
    stats = {
        'empresas_activas': empresas_activas,
        'candidatos_total': candidatos_total,
        'vacantes_activas': vacantes_activas,
        'empresas_pendientes': empresas_pendientes,
        'candidatos_pendientes': candidatos_pendientes,
        'pendientes_total': empresas_pendientes + candidatos_pendientes
    }
    return render_template("admi/inicio.html", stats=stats,
                           actividad_empresas=actividad_empresas,
                           actividad_candidatos=actividad_candidatos)

@app.route("/admin/validar-empresa")
@login_required
def admin_validar_empresa():
    if current_user.rol != "Administrador": return redirect(url_for("home"))
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT e.id_empresa, e.nombre, e.giro, e.tipo_empresa, e.telefono,
               u.correo, e.fecha_registro, d.calle, d.ciudad, d.estado,
               rh.nombre as rh_nombre, rh.correo as rh_correo
        FROM empresas e
        JOIN usuarios u ON e.id_usuario = u.id_usuario
        LEFT JOIN direcciones_empresa d ON e.id_empresa = d.id_empresa
        LEFT JOIN recursos_humanos rh ON e.id_empresa = rh.id_empresa
        WHERE e.estatus = 'pendiente'
        ORDER BY e.fecha_registro DESC
    """)
    empresas = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("admi/validar_empresa.html", empresas=empresas)

@app.route("/admin/aprobar-empresa/<int:id_empresa>", methods=["POST"])
@login_required
def admin_aprobar_empresa(id_empresa):
    if current_user.rol != "Administrador": return redirect(url_for("home"))
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("UPDATE empresas SET estatus = 'aprobada' WHERE id_empresa = %s", (id_empresa,))
    cur.execute("UPDATE usuarios SET activo = true WHERE id_usuario = (SELECT id_usuario FROM empresas WHERE id_empresa = %s)", (id_empresa,))
    cur.execute("""
        INSERT INTO validacion_empresas (id_empresa, aprobado, comentario)
        VALUES (%s, true, 'Aprobada por administrador')
    """, (id_empresa,))
    conn.commit()
    
    cur.execute("SELECT u.id_usuario, u.correo, e.nombre FROM empresas e JOIN usuarios u ON e.id_usuario = u.id_usuario WHERE e.id_empresa = %s", (id_empresa,))
    emp_data = cur.fetchone()
    if emp_data:
        id_user, correo, nombre_empresa = emp_data
        asunto = "¡Cuenta Aprobada! - Bolsa de Trabajo UT"
        cuerpo = (f"Hola {nombre_empresa},\n\nTu registro de empresa ha sido aprobado por el administrador. "
                  f"Ya puedes iniciar sesión y comenzar a publicar vacantes.\n\n"
                  f"Acceso: http://localhost:5001/login")
        crear_notificacion(id_user, "aprobacion", "Tu registro de empresa ha sido aprobado.", "/empresa", asunto, cuerpo)
        
    cur.close()
    conn.close()
    flash("Empresa aprobada exitosamente.")
    return redirect(url_for("admin_validar_empresa"))

@app.route("/admin/rechazar-empresa/<int:id_empresa>", methods=["POST"])
@login_required
def admin_rechazar_empresa(id_empresa):
    if current_user.rol != "Administrador": return redirect(url_for("home"))
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("UPDATE empresas SET estatus = 'rechazada' WHERE id_empresa = %s", (id_empresa,))
    cur.execute("""
        INSERT INTO validacion_empresas (id_empresa, aprobado, comentario)
        VALUES (%s, false, 'Rechazada por administrador')
    """, (id_empresa,))
    conn.commit()
    
    cur.execute("SELECT id_usuario FROM empresas WHERE id_empresa = %s", (id_empresa,))
    emp_user = cur.fetchone()
    if emp_user:
        asunto = "Estado de Registro - Bolsa de Trabajo UT"
        cuerpo = ("Hola,\n\nLamentamos informarte que tu registro de empresa ha sido rechazado por el administrador. "
                  "Si crees que esto es un error, por favor contacta a la coordinación.\n\n"
                  "Saludos, UT Oriental Emplea")
        crear_notificacion(emp_user[0], "rechazo", "Tu registro de empresa ha sido rechazado.", "/home", asunto, cuerpo)
        
    cur.close()
    conn.close()
    flash("Empresa rechazada.")
    return redirect(url_for("admin_validar_empresa"))

@app.route("/admin/consultar-empresas")
@login_required
def admin_consultar_empresas():
    if current_user.rol != "Administrador": return redirect(url_for("home"))
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT e.id_empresa, e.nombre, e.giro, e.tipo_empresa, e.telefono,
               u.correo, e.estatus, e.fecha_registro
        FROM empresas e
        JOIN usuarios u ON e.id_usuario = u.id_usuario
        ORDER BY e.fecha_registro DESC
    """)
    empresas = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("admi/consultar_empresas.html", empresas=empresas)

@app.route("/admin/validar-candidato")
@login_required
def admin_validar_candidato():
    if current_user.rol != "Administrador": return redirect(url_for("home"))
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT c.id_candidato, c.nombre, c.sexo, c.telefono,
               u.correo, c.fecha_registro,
               COALESCE(ca.nombre, 'Sin carrera') as carrera,
               c.anio_egreso, c.tipo_usuario
        FROM candidatos c
        JOIN usuarios u ON c.id_usuario = u.id_usuario
        LEFT JOIN carreras ca ON c.id_carrera = ca.id_carrera
        WHERE c.estatus = 'pendiente'
        ORDER BY c.fecha_registro DESC
    """)
    candidatos = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("admi/validar_candidato.html", candidatos=candidatos)

@app.route("/admin/aprobar-candidato/<int:id_candidato>", methods=["POST"])
@login_required
def admin_aprobar_candidato(id_candidato):
    if current_user.rol != "Administrador": return redirect(url_for("home"))
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("UPDATE candidatos SET estatus = 'aprobado' WHERE id_candidato = %s", (id_candidato,))
    cur.execute("UPDATE usuarios SET activo = true WHERE id_usuario = (SELECT id_usuario FROM candidatos WHERE id_candidato = %s)", (id_candidato,))
    cur.execute("""
        INSERT INTO validacion_candidatos (id_candidato, aprobado, comentario)
        VALUES (%s, true, 'Aprobado por administrador')
    """, (id_candidato,))
    conn.commit()
    
    cur.execute("SELECT u.id_usuario, u.correo, c.nombre FROM candidatos c JOIN usuarios u ON c.id_usuario = u.id_usuario WHERE c.id_candidato = %s", (id_candidato,))
    cand_data = cur.fetchone()
    if cand_data:
        id_user, correo, nombre_cand = cand_data
        asunto = "¡Cuenta Aprobada! - Bolsa de Trabajo UT"
        cuerpo = (f"Hola {nombre_cand},\n\nTu perfil profesional ha sido aprobado por el administrador. "
                  f"Ya puedes iniciar sesión y postularte a las mejores vacantes.\n\n"
                  f"Acceso: http://localhost:5001/login")
        crear_notificacion(id_user, "aprobacion", "Tu registro de candidato ha sido aprobado.", "/candidato-dashboard", asunto, cuerpo)
        
    cur.close()
    conn.close()
    flash("Candidato aprobado exitosamente.")
    return redirect(url_for("admin_validar_candidato"))

@app.route("/admin/rechazar-candidato/<int:id_candidato>", methods=["POST"])
@login_required
def admin_rechazar_candidato(id_candidato):
    if current_user.rol != "Administrador": return redirect(url_for("home"))
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("UPDATE candidatos SET estatus = 'rechazado' WHERE id_candidato = %s", (id_candidato,))
    cur.execute("""
        INSERT INTO validacion_candidatos (id_candidato, aprobado, comentario)
        VALUES (%s, false, 'Rechazado por administrador')
    """, (id_candidato,))
    conn.commit()
    
    cur.execute("SELECT id_usuario FROM candidatos WHERE id_candidato = %s", (id_candidato,))
    cand_user = cur.fetchone()
    if cand_user:
        asunto = "Estado de Registro - Bolsa de Trabajo UT"
        cuerpo = ("Hola,\n\nLamentamos informarte que tu perfil profesional ha sido rechazado por el administrador. "
                  "Para más información, puedes contactar al departamento de vinculación.\n\n"
                  "Saludos, UT Oriental Emplea")
        crear_notificacion(cand_user[0], "rechazo", "Tu registro de candidato ha sido rechazado.", "/home", asunto, cuerpo)
        
    cur.close()
    conn.close()
    flash("Candidato rechazado.")
    return redirect(url_for("admin_validar_candidato"))

@app.route("/admin/consultar-candidatos")
@login_required
def admin_consultar_candidatos():
    if current_user.rol != "Administrador": return redirect(url_for("home"))
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT c.id_candidato, c.nombre, c.sexo, c.telefono,
               u.correo, c.estatus, c.fecha_registro,
               COALESCE(ca.nombre, 'Sin carrera') as carrera,
               c.anio_egreso
        FROM candidatos c
        JOIN usuarios u ON c.id_usuario = u.id_usuario
        LEFT JOIN carreras ca ON c.id_carrera = ca.id_carrera
        ORDER BY c.fecha_registro DESC
    """)
    candidatos = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("admi/consultar_candidatos.html", candidatos=candidatos)

@app.route("/admin/vacantes")
@login_required
def admin_vacantes():
    if current_user.rol != "Administrador": return redirect(url_for("home"))
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT v.id_vacante, v.titulo, e.nombre as empresa, v.fecha_publicacion,
               v.fecha_vencimiento, v.estatus, v.salario,
               (SELECT COUNT(*) FROM postulaciones p WHERE p.id_vacante = v.id_vacante) as num_postulaciones,
               v.descripcion, v.modalidad, v.horario, v.lugar_trabajo, v.perfil,
               u.id_usuario as id_usuario_empresa
        FROM vacantes v
        JOIN empresas e ON v.id_empresa = e.id_empresa
        JOIN usuarios u ON e.id_usuario = u.id_usuario
        ORDER BY v.fecha_publicacion DESC
    """)
    vacantes = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("admi/vacantes.html", vacantes=vacantes)

@app.route("/admin/eliminar-vacante/<int:id_vacante>", methods=["POST"])
@login_required
def admin_eliminar_vacante(id_vacante):
    if current_user.rol != "Administrador":
        return redirect(url_for("home"))
    conn = get_connection()
    cur = conn.cursor()
    try:
        # Get vacancy title and the empresa's user id for notification
        cur.execute("""
            SELECT v.titulo, u.id_usuario
            FROM vacantes v
            JOIN empresas e ON v.id_empresa = e.id_empresa
            JOIN usuarios u ON e.id_usuario = u.id_usuario
            WHERE v.id_vacante = %s
        """, (id_vacante,))
        row = cur.fetchone()
        if row:
            titulo_vacante = row[0]
            id_usuario_empresa = row[1]
            # Delete related records first
            cur.execute("DELETE FROM postulaciones WHERE id_vacante = %s", (id_vacante,))
            cur.execute("DELETE FROM requisitos_vacante WHERE id_vacante = %s", (id_vacante,))
            cur.execute("DELETE FROM vacantes_guardadas WHERE id_vacante = %s", (id_vacante,))
            cur.execute("DELETE FROM vacantes WHERE id_vacante = %s", (id_vacante,))
            conn.commit()
            # Notify the empresa
            crear_notificacion(
                id_usuario_empresa,
                "vacantes",
                f"Tu vacante '{titulo_vacante}' fue eliminada por un administrador.",
                "/empresa/ver-vacantes"
            )
            flash(f"Vacante '{titulo_vacante}' eliminada correctamente.")
        else:
            flash("La vacante no fue encontrada.")
    except Exception as e:
        conn.rollback()
        print(f"Error al eliminar vacante: {e}")
        flash("Error al eliminar la vacante.")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("admin_vacantes"))

@app.route("/admin/seguimiento")
@login_required
def admin_seguimiento():
    if current_user.rol != "Administrador": return redirect(url_for("home"))
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT p.id_postulacion, v.titulo, e.nombre as empresa,
               c.nombre as candidato, p.estado, p.fecha_postulacion,
               v.salario, v.descripcion, v.modalidad,
               ue.correo as correo_empresa, uc.correo as correo_candidato,
               v.lugar_trabajo
        FROM postulaciones p
        JOIN vacantes v ON p.id_vacante = v.id_vacante
        JOIN empresas e ON v.id_empresa = e.id_empresa
        JOIN candidatos c ON p.id_candidato = c.id_candidato
        JOIN usuarios ue ON e.id_usuario = ue.id_usuario
        JOIN usuarios uc ON c.id_usuario = uc.id_usuario
        ORDER BY p.fecha_postulacion DESC
        LIMIT 50
    """)
    seguimiento = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("admi/seguimiento.html", seguimiento=seguimiento)

@app.route("/admin/reportes")
@login_required
def admin_reportes():
    if current_user.rol != "Administrador": return redirect(url_for("home"))
    
    fecha_inicio = request.args.get("fecha_inicio")
    fecha_fin = request.args.get("fecha_fin")
    
    conn = get_connection()
    cur = conn.cursor()
    
    # Construir cláusulas WHERE dinámicas
    we = wc = wv = wp = ""
    we_a = wc_a = wv_a = wp_a = "" # para queries con alias e, c, v, p
    params = ()
    
    if fecha_inicio and fecha_fin:
        dt_cond = " >= %s AND {} <= %s::date + INTERVAL '1 day' - INTERVAL '1 second'"
        we = "WHERE fecha_registro" + dt_cond.format("fecha_registro")
        wc = "WHERE fecha_registro" + dt_cond.format("fecha_registro")
        wv = "WHERE fecha_publicacion" + dt_cond.format("fecha_publicacion")
        wp = "WHERE fecha_postulacion" + dt_cond.format("fecha_postulacion")
        
        we_a = "WHERE e.fecha_registro" + dt_cond.format("e.fecha_registro")
        wc_a = "WHERE c.fecha_registro" + dt_cond.format("c.fecha_registro")
        wv_a = "WHERE v.fecha_publicacion" + dt_cond.format("v.fecha_publicacion")
        wp_a = "WHERE p.fecha_postulacion" + dt_cond.format("p.fecha_postulacion")
        params = (fecha_inicio, fecha_fin)

    # Helper para queries
    def get_count(query, p=()):
        cur.execute(query, p)
        res = cur.fetchone()
        return res[0] if res else 0

    def get_group(query, p=()):
        cur.execute(query, p)
        return {row[0]: row[1] for row in cur.fetchall()}

    # Totales
    total_empresas = get_count(f"SELECT COUNT(*) FROM empresas {we}", params)
    total_candidatos = get_count(f"SELECT COUNT(*) FROM candidatos {wc}", params)
    total_vacantes = get_count(f"SELECT COUNT(*) FROM vacantes {wv}", params)
    total_postulaciones = get_count(f"SELECT COUNT(*) FROM postulaciones {wp}", params)

    # Estatus
    empresas_estatus = get_group(f"SELECT estatus, COUNT(*) FROM empresas {we} GROUP BY estatus", params)
    vacantes_estatus = get_group(f"SELECT estatus, COUNT(*) FROM vacantes {wv} GROUP BY estatus", params)
    postulaciones_estado = get_group(f"SELECT estado, COUNT(*) FROM postulaciones {wp} GROUP BY estado", params)

    # Top 5 empresas con más vacantes
    cur.execute(f"""
        SELECT e.nombre, COUNT(v.id_vacante) as total
        FROM empresas e
        LEFT JOIN vacantes v ON e.id_empresa = v.id_empresa
        {we_a}
        GROUP BY e.nombre ORDER BY total DESC LIMIT 5
    """, params)
    top_empresas = cur.fetchall()

    # Candidatos por carrera
    cur.execute(f"""
        SELECT COALESCE(ca.nombre, 'Sin carrera'), COUNT(c.id_candidato)
        FROM candidatos c
        LEFT JOIN carreras ca ON c.id_carrera = ca.id_carrera
        {wc_a}
        GROUP BY ca.nombre ORDER BY COUNT(c.id_candidato) DESC LIMIT 8
    """, params)
    candidatos_carrera = cur.fetchall()

    # Candidatos por sexo
    candidatos_sexo = get_group(f"SELECT COALESCE(sexo, 'No especificado'), COUNT(*) FROM candidatos c {wc_a} GROUP BY sexo", params)

    # Top 5 vacantes con más postulaciones
    cur.execute(f"""
        SELECT v.titulo, COUNT(p.id_postulacion) as total
        FROM vacantes v
        LEFT JOIN postulaciones p ON v.id_vacante = p.id_vacante
        {wv_a}
        GROUP BY v.titulo ORDER BY total DESC LIMIT 5
    """, params)
    top_vacantes_postuladas = cur.fetchall()

    # Registros por mes (últimos 6 meses). Este no se filtra por la fecha seleccionada
    # ya que es una vista histórica de 6 meses fija.
    cur.execute("""
        SELECT TO_CHAR(fecha_registro, 'YYYY-MM') as mes, COUNT(*)
        FROM candidatos
        WHERE fecha_registro >= NOW() - INTERVAL '6 months'
        GROUP BY mes ORDER BY mes
    """)
    registros_mes = cur.fetchall()

    # Tasa de aceptación
    total_aceptados = get_count(f"SELECT COUNT(*) FROM postulaciones p {wp_a} {'AND' if wp_a else 'WHERE'} p.estado = 'aceptado'", params)
    tasa_aceptacion = round((total_aceptados / total_postulaciones * 100), 1) if total_postulaciones > 0 else 0

    # Reporte de Colocación (Nuevo)
    query_colocacion = """
        SELECT 
            TO_CHAR(MAX(v.fecha_publicacion), 'DD/MM/YYYY') as fecha_publicacion,
            e.nombre as empresa,
            'Conv-' || EXTRACT(YEAR FROM MIN(e.fecha_registro)) || '-' || LPAD(e.id_empresa::text, 3, '0') as convenio,
            COUNT(DISTINCT v.id_vacante) as espacios_gestionados,
            COUNT(DISTINCT CASE WHEN c.sexo = 'Masculino' THEN p.id_postulacion END) as postulados_h,
            COUNT(DISTINCT CASE WHEN c.sexo = 'Femenino' THEN p.id_postulacion END) as postulados_m,
            COUNT(DISTINCT p.id_postulacion) as total_postulados,
            COUNT(DISTINCT CASE WHEN c.sexo = 'Masculino' AND p.estado = 'aceptado' THEN p.id_postulacion END) as colocados_h,
            COUNT(DISTINCT CASE WHEN c.sexo = 'Femenino' AND p.estado = 'aceptado' THEN p.id_postulacion END) as colocados_m,
            COUNT(DISTINCT CASE WHEN p.estado = 'aceptado' THEN p.id_postulacion END) as total_colocados
        FROM empresas e
        LEFT JOIN vacantes v ON e.id_empresa = v.id_empresa
        LEFT JOIN postulaciones p ON v.id_vacante = p.id_vacante
        LEFT JOIN candidatos c ON p.id_candidato = c.id_candidato
        WHERE e.estatus = 'aprobada'
    """
    if fecha_inicio and fecha_fin:
        query_colocacion += " AND v.fecha_publicacion >= %s AND v.fecha_publicacion <= %s::date + INTERVAL '1 day' - INTERVAL '1 second' "
        
    query_colocacion += """
        GROUP BY e.id_empresa, e.nombre
        HAVING COUNT(DISTINCT v.id_vacante) > 0
        ORDER BY MAX(v.fecha_publicacion) DESC NULLS LAST
    """
    cur.execute(query_colocacion, params if fecha_inicio and fecha_fin else ())
    reporte_colocacion = cur.fetchall()

    cur.close()
    conn.close()

    reportes = {
        'total_empresas': total_empresas,
        'total_candidatos': total_candidatos,
        'total_vacantes': total_vacantes,
        'total_postulaciones': total_postulaciones,
        'empresas_aprobadas': empresas_estatus.get('aprobada', 0),
        'empresas_pendientes': empresas_estatus.get('pendiente', 0),
        'empresas_rechazadas': empresas_estatus.get('rechazada', 0),
        'vacantes_activas': vacantes_estatus.get('activa', 0),
        'vacantes_inactivas': vacantes_estatus.get('inactiva', 0) + vacantes_estatus.get('vencida', 0),
        'post_pendientes': postulaciones_estado.get('pendiente', 0),
        'post_aceptados': postulaciones_estado.get('aceptado', 0),
        'post_rechazados': postulaciones_estado.get('rechazado', 0),
        'top_empresas_nombres': [r[0] for r in top_empresas],
        'top_empresas_vacantes': [r[1] for r in top_empresas],
        'candidatos_carrera_nombres': [r[0] for r in candidatos_carrera],
        'candidatos_carrera_totales': [r[1] for r in candidatos_carrera],
        'candidatos_masculino': candidatos_sexo.get('Masculino', 0),
        'candidatos_femenino': candidatos_sexo.get('Femenino', 0),
        'candidatos_otro': candidatos_sexo.get('Otro', 0) + candidatos_sexo.get('No especificado', 0),
        'top_vacantes_nombres': [r[0] for r in top_vacantes_postuladas],
        'top_vacantes_postulaciones': [r[1] for r in top_vacantes_postuladas],
        'registros_mes_labels': [r[0] for r in registros_mes],
        'registros_mes_totales': [r[1] for r in registros_mes],
        'tasa_aceptacion': tasa_aceptacion,
        'fecha_inicio': fecha_inicio or '',
        'fecha_fin': fecha_fin or '',
        'colocacion': reporte_colocacion
    }
    return render_template("admi/reportes.html", reportes=reportes)

@app.route("/admin/reportes/exportar-excel")
@login_required
def admin_exportar_excel():
    if current_user.rol != "Administrador": return redirect(url_for("home"))
    
    fecha_inicio = request.args.get("fecha_inicio")
    fecha_fin = request.args.get("fecha_fin")
    
    # Construir cláusulas WHERE dinámicas
    we = wc = wv = wp = ""
    we_a = wc_a = wv_a = wp_a = "" # con alias e, c, v, p
    params = ()
    
    if fecha_inicio and fecha_fin:
        dt_cond = " >= %s AND {} <= %s::date + INTERVAL '1 day' - INTERVAL '1 second'"
        we = "WHERE fecha_registro" + dt_cond.format("fecha_registro")
        wc = "WHERE fecha_registro" + dt_cond.format("fecha_registro")
        wv = "WHERE fecha_publicacion" + dt_cond.format("fecha_publicacion")
        wp = "WHERE fecha_postulacion" + dt_cond.format("fecha_postulacion")
        
        we_a = "WHERE e.fecha_registro" + dt_cond.format("e.fecha_registro")
        wc_a = "WHERE c.fecha_registro" + dt_cond.format("c.fecha_registro")
        wv_a = "WHERE v.fecha_publicacion" + dt_cond.format("v.fecha_publicacion")
        wp_a = "WHERE p.fecha_postulacion" + dt_cond.format("p.fecha_postulacion")
        params = (fecha_inicio, fecha_fin)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    
    wb = Workbook()
    
    # Colores institucionales
    verde_fill = PatternFill(start_color="0E312D", end_color="0E312D", fill_type="solid")
    dorado_fill = PatternFill(start_color="C2914F", end_color="C2914F", fill_type="solid")
    gris_claro_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    blanco_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    font_header = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    font_title = Font(name="Calibri", bold=True, color="0E312D", size=18)
    font_subtitle = Font(name="Calibri", bold=True, color="C2914F", size=13)
    font_normal = Font(name="Calibri", size=11, color="333333")
    font_bold = Font(name="Calibri", bold=True, size=11, color="1A202C")
    font_number = Font(name="Calibri", bold=True, size=14, color="0E312D")
    
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0")
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    def auto_width(ws):
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 4, 45)
    
    def style_header_row(ws, row_num, num_cols):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = verde_fill
            cell.font = font_header
            cell.alignment = center_align
            cell.border = thin_border
    
    def style_data_rows(ws, start_row, end_row, num_cols):
        for row in range(start_row, end_row + 1):
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = font_normal
                cell.alignment = left_align
                cell.border = thin_border
                if row % 2 == 0:
                    cell.fill = gris_claro_fill
    
    conn = get_connection()
    cur = conn.cursor()
    
    # ===== HOJA 1: RESUMEN EJECUTIVO =====
    ws1 = wb.active
    ws1.title = "Resumen Ejecutivo"
    ws1.sheet_properties.tabColor = "0E312D"
    
    ws1.merge_cells("A1:D1")
    ws1["A1"].value = "BOLSA DE TRABAJO - UT ORIENTAL"
    ws1["A1"].font = font_title
    ws1["A1"].alignment = center_align
    
    ws1.merge_cells("A2:D2")
    ws1["A2"].value = "Reporte Consolidado del Sistema"
    ws1["A2"].font = font_subtitle
    ws1["A2"].alignment = center_align
    
    ws1.merge_cells("A3:D3")
    from datetime import datetime as dt_now
    ws1["A3"].value = f"Fecha de generación: {dt_now.now().strftime('%d/%m/%Y %H:%M')}"
    if fecha_inicio and fecha_fin:
        ws1["A3"].value += f" | Filtro: {fecha_inicio} al {fecha_fin}"
    ws1["A3"].font = Font(name="Calibri", italic=True, color="666666", size=10)
    ws1["A3"].alignment = center_align
    
    # Helper functions for getting counts
    def get_count_xls(query, p=()):
        cur.execute(query, p)
        res = cur.fetchone()
        return res[0] if res else 0

    # Totales
    total_empresas = get_count_xls(f"SELECT COUNT(*) FROM empresas {we}", params)
    total_candidatos = get_count_xls(f"SELECT COUNT(*) FROM candidatos {wc}", params)
    total_vacantes = get_count_xls(f"SELECT COUNT(*) FROM vacantes {wv}", params)
    total_postulaciones = get_count_xls(f"SELECT COUNT(*) FROM postulaciones {wp}", params)
    
    row = 5
    metrics = [
        ("Total de Empresas", total_empresas),
        ("Total de Candidatos", total_candidatos),
        ("Total de Vacantes", total_vacantes),
        ("Total de Postulaciones", total_postulaciones),
    ]
    
    ws1.cell(row=row, column=1, value="INDICADOR").fill = dorado_fill
    ws1.cell(row=row, column=1).font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    ws1.cell(row=row, column=1).alignment = center_align
    ws1.cell(row=row, column=1).border = thin_border
    ws1.cell(row=row, column=2, value="TOTAL").fill = dorado_fill
    ws1.cell(row=row, column=2).font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    ws1.cell(row=row, column=2).alignment = center_align
    ws1.cell(row=row, column=2).border = thin_border
    
    for i, (label, value) in enumerate(metrics):
        r = row + 1 + i
        ws1.cell(row=r, column=1, value=label).font = font_bold
        ws1.cell(row=r, column=1).alignment = left_align
        ws1.cell(row=r, column=1).border = thin_border
        ws1.cell(row=r, column=2, value=value).font = font_number
        ws1.cell(row=r, column=2).alignment = center_align
        ws1.cell(row=r, column=2).border = thin_border
        if i % 2 == 1:
            ws1.cell(row=r, column=1).fill = gris_claro_fill
            ws1.cell(row=r, column=2).fill = gris_claro_fill
    
    # Empresas por estatus
    row = 11
    ws1.cell(row=row, column=1, value="EMPRESAS POR ESTATUS").font = font_subtitle
    ws1.merge_cells(f"A{row}:B{row}")
    row += 1
    cur.execute(f"SELECT estatus, COUNT(*) FROM empresas {we} GROUP BY estatus", params)
    for estatus, count in cur.fetchall():
        ws1.cell(row=row, column=1, value=estatus.capitalize()).font = font_normal
        ws1.cell(row=row, column=1).border = thin_border
        ws1.cell(row=row, column=2, value=count).font = font_bold
        ws1.cell(row=row, column=2).alignment = center_align
        ws1.cell(row=row, column=2).border = thin_border
        row += 1
    
    # Vacantes por estatus
    row += 1
    ws1.cell(row=row, column=1, value="VACANTES POR ESTATUS").font = font_subtitle
    ws1.merge_cells(f"A{row}:B{row}")
    row += 1
    cur.execute(f"SELECT estatus, COUNT(*) FROM vacantes {wv} GROUP BY estatus", params)
    for estatus, count in cur.fetchall():
        ws1.cell(row=row, column=1, value=estatus.capitalize()).font = font_normal
        ws1.cell(row=row, column=1).border = thin_border
        ws1.cell(row=row, column=2, value=count).font = font_bold
        ws1.cell(row=row, column=2).alignment = center_align
        ws1.cell(row=row, column=2).border = thin_border
        row += 1
    
    # Postulaciones por estado
    row += 1
    ws1.cell(row=row, column=1, value="POSTULACIONES POR ESTADO").font = font_subtitle
    ws1.merge_cells(f"A{row}:B{row}")
    row += 1
    cur.execute(f"SELECT estado, COUNT(*) FROM postulaciones {wp} GROUP BY estado", params)
    for estado, count in cur.fetchall():
        ws1.cell(row=row, column=1, value=estado.capitalize()).font = font_normal
        ws1.cell(row=row, column=1).border = thin_border
        ws1.cell(row=row, column=2, value=count).font = font_bold
        ws1.cell(row=row, column=2).alignment = center_align
        ws1.cell(row=row, column=2).border = thin_border
        row += 1
    
    auto_width(ws1)
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 18
    
    # ===== HOJA 2: EMPRESAS =====
    ws2 = wb.create_sheet("Empresas")
    ws2.sheet_properties.tabColor = "C2914F"
    headers_emp = ["ID", "Nombre", "Giro", "Tipo", "Teléfono", "Correo", "Estatus", "Fecha Registro"]
    for i, h in enumerate(headers_emp, 1):
        ws2.cell(row=1, column=i, value=h)
    style_header_row(ws2, 1, len(headers_emp))
    
    cur.execute(f"""
        SELECT e.id_empresa, e.nombre, e.giro, e.tipo_empresa, e.telefono,
               u.correo, e.estatus, e.fecha_registro
        FROM empresas e
        JOIN usuarios u ON e.id_usuario = u.id_usuario
        {we_a}
        ORDER BY e.fecha_registro DESC
    """, params)
    empresas = cur.fetchall()
    for r, emp in enumerate(empresas, 2):
        for c, val in enumerate(emp, 1):
            cell = ws2.cell(row=r, column=c, value=str(val) if val else "")
    style_data_rows(ws2, 2, len(empresas) + 1, len(headers_emp))
    auto_width(ws2)
    
    # ===== HOJA 3: CANDIDATOS =====
    ws3 = wb.create_sheet("Candidatos")
    ws3.sheet_properties.tabColor = "0E312D"
    headers_cand = ["ID", "Nombre", "Apellido Pat.", "Apellido Mat.", "Sexo", "Teléfono", "Correo", "Carrera", "Año Egreso", "Estatus", "Idiomas", "Cursos y Cert.", "Fecha Registro"]
    for i, h in enumerate(headers_cand, 1):
        ws3.cell(row=1, column=i, value=h)
    style_header_row(ws3, 1, len(headers_cand))
    
    cur.execute(f"""
        SELECT c.id_candidato, c.nombre, c.apellido_paterno, c.apellido_materno,
               c.sexo, c.telefono, u.correo,
               c.carrera,
               c.anio_egreso, c.estatus, c.idiomas, c.cursos_certificaciones, c.fecha_registro
        FROM candidatos c
        JOIN usuarios u ON c.id_usuario = u.id_usuario
        {wc_a}
        ORDER BY c.fecha_registro DESC
    """, params)
    candidatos = cur.fetchall()
    for r, cand in enumerate(candidatos, 2):
        for c, val in enumerate(cand, 1):
            ws3.cell(row=r, column=c, value=str(val) if val else "")
    style_data_rows(ws3, 2, len(candidatos) + 1, len(headers_cand))
    auto_width(ws3)
    
    # ===== HOJA 4: VACANTES =====
    ws4 = wb.create_sheet("Vacantes")
    ws4.sheet_properties.tabColor = "C2914F"
    headers_vac = ["ID", "Título", "Empresa", "Salario", "Modalidad", "Horario", "Lugar", "Fecha Publicación", "Estatus", "Postulaciones"]
    for i, h in enumerate(headers_vac, 1):
        ws4.cell(row=1, column=i, value=h)
    style_header_row(ws4, 1, len(headers_vac))
    
    cur.execute(f"""
        SELECT v.id_vacante, v.titulo, e.nombre, v.salario, v.modalidad,
               v.horario, v.lugar_trabajo, v.fecha_publicacion, v.estatus,
               (SELECT COUNT(*) FROM postulaciones p WHERE p.id_vacante = v.id_vacante)
        FROM vacantes v
        JOIN empresas e ON v.id_empresa = e.id_empresa
        {wv_a}
        ORDER BY v.fecha_publicacion DESC
    """, params)
    vacantes = cur.fetchall()
    for r, vac in enumerate(vacantes, 2):
        for c, val in enumerate(vac, 1):
            if c == 4 and val:  # Salario
                ws4.cell(row=r, column=c, value=float(val))
                ws4.cell(row=r, column=c).number_format = '$#,##0.00'
            else:
                ws4.cell(row=r, column=c, value=str(val) if val else "")
    style_data_rows(ws4, 2, len(vacantes) + 1, len(headers_vac))
    auto_width(ws4)
    
    # ===== HOJA 5: POSTULACIONES =====
    ws5 = wb.create_sheet("Postulaciones")
    ws5.sheet_properties.tabColor = "0E312D"
    headers_post = ["ID", "Vacante", "Empresa", "Candidato", "Estado", "Fecha Postulación", "Comentarios"]
    for i, h in enumerate(headers_post, 1):
        ws5.cell(row=1, column=i, value=h)
    style_header_row(ws5, 1, len(headers_post))
    
    cur.execute(f"""
        SELECT p.id_postulacion, v.titulo, e.nombre, c.nombre,
               p.estado, p.fecha_postulacion, COALESCE(p.comentarios, '')
        FROM postulaciones p
        JOIN vacantes v ON p.id_vacante = v.id_vacante
        JOIN empresas e ON v.id_empresa = e.id_empresa
        JOIN candidatos c ON p.id_candidato = c.id_candidato
        {wp_a}
        ORDER BY p.fecha_postulacion DESC
    """, params)
    postulaciones = cur.fetchall()
    for r, post in enumerate(postulaciones, 2):
        for c, val in enumerate(post, 1):
            ws5.cell(row=r, column=c, value=str(val) if val else "")
    style_data_rows(ws5, 2, len(postulaciones) + 1, len(headers_post))
    auto_width(ws5)
    
    # ===== HOJA 6: ESTADÍSTICAS =====
    ws6 = wb.create_sheet("Estadísticas")
    ws6.sheet_properties.tabColor = "C2914F"
    
    ws6.merge_cells("A1:C1")
    ws6["A1"].value = "TOP 5 EMPRESAS CON MÁS VACANTES"
    ws6["A1"].font = font_subtitle
    
    ws6.cell(row=2, column=1, value="Empresa").fill = verde_fill
    ws6.cell(row=2, column=1).font = font_header
    ws6.cell(row=2, column=1).border = thin_border
    ws6.cell(row=2, column=2, value="Vacantes").fill = verde_fill
    ws6.cell(row=2, column=2).font = font_header
    ws6.cell(row=2, column=2).border = thin_border
    
    cur.execute(f"""
        SELECT e.nombre, COUNT(v.id_vacante) as total
        FROM empresas e
        LEFT JOIN vacantes v ON e.id_empresa = v.id_empresa
        {we_a}
        GROUP BY e.nombre ORDER BY total DESC LIMIT 5
    """, params)
    top_emp = cur.fetchall()
    for i, (nombre, total) in enumerate(top_emp, 3):
        ws6.cell(row=i, column=1, value=nombre).font = font_normal
        ws6.cell(row=i, column=1).border = thin_border
        ws6.cell(row=i, column=2, value=total).font = font_bold
        ws6.cell(row=i, column=2).alignment = center_align
        ws6.cell(row=i, column=2).border = thin_border
    
    row = len(top_emp) + 5
    ws6.merge_cells(f"A{row}:C{row}")
    ws6[f"A{row}"].value = "CANDIDATOS POR CARRERA"
    ws6[f"A{row}"].font = font_subtitle
    
    row += 1
    ws6.cell(row=row, column=1, value="Carrera").fill = verde_fill
    ws6.cell(row=row, column=1).font = font_header
    ws6.cell(row=row, column=1).border = thin_border
    ws6.cell(row=row, column=2, value="Candidatos").fill = verde_fill
    ws6.cell(row=row, column=2).font = font_header
    ws6.cell(row=row, column=2).border = thin_border
    
    cur.execute(f"""
        SELECT COALESCE(c.carrera, 'Sin carrera'), COUNT(c.id_candidato)
        FROM candidatos c
        {wc_a}
        GROUP BY ca.nombre ORDER BY COUNT(c.id_candidato) DESC
    """, params)
    carreras_data = cur.fetchall()
    for i, (carrera, count) in enumerate(carreras_data, row + 1):
        ws6.cell(row=i, column=1, value=carrera).font = font_normal
        ws6.cell(row=i, column=1).border = thin_border
        ws6.cell(row=i, column=2, value=count).font = font_bold
        ws6.cell(row=i, column=2).alignment = center_align
        ws6.cell(row=i, column=2).border = thin_border
    
    auto_width(ws6)
    ws6.column_dimensions['A'].width = 35
    
    cur.close()
    conn.close()
    
    # Generar respuesta
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=Reporte_UT_Emplea.xlsx'
    return response

@app.route("/admin/reportes/exportar-colocacion-excel")
@login_required
def admin_exportar_colocacion_excel():
    if current_user.rol != "Administrador": return redirect(url_for("home"))
    
    fecha_inicio = request.args.get("fecha_inicio")
    fecha_fin = request.args.get("fecha_fin")
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Colocación"
    ws.sheet_properties.tabColor = "0E312D"
    
    # Estilos
    verde_fill = PatternFill(start_color="0E312D", end_color="0E312D", fill_type="solid")
    blanco_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    gris_claro_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    dorado_fondo = PatternFill(start_color="FCEFDC", end_color="FCEFDC", fill_type="solid")
    
    font_header = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    font_title = Font(name="Calibri", bold=True, color="0E312D", size=18)
    font_subtitle = Font(name="Calibri", bold=True, color="C2914F", size=13)
    font_normal = Font(name="Calibri", size=11, color="475569")
    font_bold = Font(name="Calibri", bold=True, size=11, color="1E293B")
    font_red = Font(name="Calibri", bold=True, size=11, color="DC2626")
    font_green = Font(name="Calibri", bold=True, size=11, color="16A34A")
    
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0")
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Encabezado principal del archivo
    ws.merge_cells("A1:J1")
    ws["A1"].value = "BOLSA DE TRABAJO - UT ORIENTAL"
    ws["A1"].font = font_title
    ws["A1"].alignment = center_align
    
    ws.merge_cells("A2:J2")
    ws["A2"].value = "Reporte de Colocación y Espacios Gestionados por Convenio"
    ws["A2"].font = font_subtitle
    ws["A2"].alignment = center_align
    
    ws.merge_cells("A3:J3")
    from datetime import datetime as dt_now
    ws["A3"].value = f"Fecha de generación: {dt_now.now().strftime('%d/%m/%Y %H:%M')}"
    if fecha_inicio and fecha_fin:
        ws["A3"].value += f" | Filtro temporal aplicado: {fecha_inicio} al {fecha_fin}"
    ws["A3"].font = Font(name="Calibri", italic=True, color="64748B", size=10)
    ws["A3"].alignment = center_align
    
    # Encabezados de tabla
    headers = [
        "Fecha de Publicación", "Empresa", "Convenio", "Espacios Gestionados", 
        "Postulados (H)", "Postulados (M)", "Total Postulados", 
        "Colocados (H)", "Colocados (M)", "Total Colocados"
    ]
    
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=i, value=h)
        cell.fill = verde_fill
        cell.font = font_header
        cell.alignment = center_align
        cell.border = thin_border
        
    # Consulta a BD
    conn = get_connection()
    cur = conn.cursor()
    params = ()
    query_colocacion = """
        SELECT 
            TO_CHAR(MAX(v.fecha_publicacion), 'DD/MM/YYYY') as fecha_publicacion,
            e.nombre as empresa,
            'Conv-' || EXTRACT(YEAR FROM MIN(e.fecha_registro)) || '-' || LPAD(e.id_empresa::text, 3, '0') as convenio,
            COUNT(DISTINCT v.id_vacante) as espacios_gestionados,
            COUNT(DISTINCT CASE WHEN c.sexo = 'Masculino' THEN p.id_postulacion END) as postulados_h,
            COUNT(DISTINCT CASE WHEN c.sexo = 'Femenino' THEN p.id_postulacion END) as postulados_m,
            COUNT(DISTINCT p.id_postulacion) as total_postulados,
            COUNT(DISTINCT CASE WHEN c.sexo = 'Masculino' AND p.estado = 'aceptado' THEN p.id_postulacion END) as colocados_h,
            COUNT(DISTINCT CASE WHEN c.sexo = 'Femenino' AND p.estado = 'aceptado' THEN p.id_postulacion END) as colocados_m,
            COUNT(DISTINCT CASE WHEN p.estado = 'aceptado' THEN p.id_postulacion END) as total_colocados
        FROM empresas e
        LEFT JOIN vacantes v ON e.id_empresa = v.id_empresa
        LEFT JOIN postulaciones p ON v.id_vacante = p.id_vacante
        LEFT JOIN candidatos c ON p.id_candidato = c.id_candidato
        WHERE e.estatus = 'aprobada'
    """
    if fecha_inicio and fecha_fin:
        query_colocacion += " AND v.fecha_publicacion >= %s AND v.fecha_publicacion <= %s::date + INTERVAL '1 day' - INTERVAL '1 second' "
        params = (fecha_inicio, fecha_fin)
        
    query_colocacion += """
        GROUP BY e.id_empresa, e.nombre
        HAVING COUNT(DISTINCT v.id_vacante) > 0
        ORDER BY MAX(v.fecha_publicacion) DESC NULLS LAST
    """
    
    cur.execute(query_colocacion, params)
    filas = cur.fetchall()
    
    # Llenado de filas
    for r, fila in enumerate(filas, 6):
        for c, val in enumerate(fila, 1):
            cell = ws.cell(row=r, column=c, value=str(val) if val else "0")
            cell.border = thin_border
            
            # Formateo visual específico por columna igual a la vista HTML
            if c == 1: # Fecha
                cell.alignment = left_align
                cell.font = font_normal
            elif c == 2: # Empresa
                cell.alignment = left_align
                cell.font = font_bold
            elif c == 3: # Convenio
                cell.alignment = center_align
                cell.font = Font(name="Calibri", size=10, color="64748B")
                cell.fill = gris_claro_fill
            elif c == 4 or c == 7: # Espacios o Total Postulados
                cell.alignment = center_align
                cell.font = font_red
            elif c == 10: # Total Colocados
                cell.alignment = center_align
                cell.font = font_green
            else:
                cell.alignment = center_align
                cell.font = font_normal
                
        if r % 2 == 0: # Cebra en las filas restando columnas con diseño especial
            for c in [1, 2, 4, 5, 6, 7, 8, 9, 10]:
                ws.cell(row=r, column=c).fill = gris_claro_fill
    
    if len(filas) == 0:
        ws.merge_cells("A6:J6")
        cell = ws["A6"]
        cell.value = "No hay datos de empresas con vacantes en este rango temporal."
        cell.alignment = center_align
        cell.font = font_normal
        for i in range(1, 11):
            ws.cell(row=6, column=i).border = thin_border
                
    cur.close()
    conn.close()
    
    # Autoajuste de columnas
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.row > 4 and cell.value: # Saltar encabezados gigantes
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 6, 45)
        
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=Reporte_Colocacion_UT_Emplea.xlsx'
    return response

# ================= EMPRESA REGISTRO =================
@app.route("/registro-empresa", methods=["GET", "POST"])
def registro_empresa():
    if request.method == "POST":
        nombre_empresa = strip_tags(request.form.get("empresa"))
        giro = strip_tags(request.form.get("giro"))
        tipo_empresa = strip_tags(request.form.get("tipo_empresa"))
        telefono = strip_tags(request.form.get("telefono"))
        direccion = strip_tags(request.form.get("direccion"))
        correo = request.form.get("correo")
        password = request.form.get("password")
        responsable_rrhh = strip_tags(request.form.get("responsable_rrhh"))
        telefono_rrhh = strip_tags(request.form.get("telefono_rrhh"))
        correo_rrhh = strip_tags(request.form.get("correo_rrhh"))
        codigo_postal = strip_tags(request.form.get("codigo_postal"))

        # Validaciones
        error = validar_registro_datos(correo, password, telefono, codigo_postal)
        if not error:
            # Detección de SQL Injection y Scripts en campos de texto
            for campo in [nombre_empresa, giro, direccion, responsable_rrhh]:
                if detectar_script(campo) or detectar_sqli(campo):
                    error = "⚠️ Seguridad: Se detectó contenido malicioso en los campos. Solicitud bloqueada."
                    break
        
        if error:
            flash(error)
            return redirect(url_for("registro_empresa"))

        hashed_password = generate_password_hash(password)
        conn = get_connection()
        cur = conn.cursor()
        try:
            cur.execute("""
                INSERT INTO usuarios (id_rol, correo, password, activo) 
                VALUES (2, %s, %s, false) RETURNING id_usuario
            """, (correo, hashed_password))
            id_usuario = cur.fetchone()[0]
            cur.execute("""
                INSERT INTO empresas (id_usuario, nombre, giro, tipo_empresa, telefono, estatus) 
                VALUES (%s, %s, %s, %s, %s, 'pendiente') RETURNING id_empresa
            """, (id_usuario, nombre_empresa, giro, tipo_empresa, telefono))
            id_empresa = cur.fetchone()[0]
            cur.execute("""
                INSERT INTO direcciones_empresa (id_empresa, calle, codigo_postal) VALUES (%s, %s, %s)
            """, (id_empresa, direccion, codigo_postal))
            cur.execute("""
                INSERT INTO recursos_humanos (id_empresa, nombre, telefono, correo) 
                VALUES (%s, %s, %s, %s)
            """, (id_empresa, responsable_rrhh, telefono_rrhh, correo_rrhh))
            conn.commit()
            notificar_admins("registro", f"Nueva empresa registrada: {nombre_empresa}", "/admin/validar-empresa")
            flash("Empresa registrada exitosamente. Su cuenta será validada por un administrador.")
            return redirect(url_for("home"))
        except Exception as e:
            conn.rollback()
            flash(f"Error al registrar: {str(e)}")
            return redirect(url_for("registro_empresa"))
        finally:
            cur.close()
            conn.close()
    return render_template("empresa/registro-empresa.html")

# ================= CANDIDATO =================
@app.route("/registro-candidato", methods=["GET", "POST"])
def registro_candidato():
    if request.method == "POST":
        # Datos de Usuario
        correo = request.form.get("correo")
        password = request.form.get("password")
        
        # Datos Personales
        nombre = strip_tags(request.form.get("nombre"))
        apellido_paterno = strip_tags(request.form.get("apellido_paterno"))
        apellido_materno = strip_tags(request.form.get("apellido_materno"))
        sexo = strip_tags(request.form.get("sexo"))
        telefono = strip_tags(request.form.get("telefono"))
        codigo_postal = strip_tags(request.form.get("codigo_postal"))
        ubicacion = strip_tags(request.form.get("ubicacion"))
        
        # Validaciones
        error = validar_registro_datos(correo, password, telefono, codigo_postal)
        if error:
            flash(error)
            return redirect(url_for("registro_candidato"))
        
        # Datos Profesionales (Educación)
        carrera = strip_tags(request.form.get("carrera"))
        
        anio_egreso = request.form.get("egreso")
        anio_egreso = int(anio_egreso) if anio_egreso and anio_egreso.isdigit() else None
        
        ultimo_grado_estudios = strip_tags(request.form.get("ultimo_grado_estudios"))
        institucion_estudios = strip_tags(request.form.get("institucion_estudios"))
        carrera_estudios = strip_tags(request.form.get("carrera_estudios"))
        
        # Datos Laborales
        ultimo_puesto = strip_tags(request.form.get("ultimo_puesto"))
        ultima_empresa = strip_tags(request.form.get("ultima_empresa"))
        fecha_inicio_laboral = strip_tags(request.form.get("fecha_inicio_laboral"))
        fecha_fin_laboral = strip_tags(request.form.get("fecha_fin_laboral"))
        actividades_logros = strip_tags(request.form.get("actividades_logros"))
        
        # Intereses
        puesto_deseado = strip_tags(request.form.get("puesto_deseado"))
        sueldo_deseado = request.form.get("sueldo_deseado")
        sueldo_deseado = float(sueldo_deseado) if sueldo_deseado and sueldo_deseado.strip() else 0
        
        area_trabajo = strip_tags(request.form.get("area_trabajo"))
        especialidad = strip_tags(request.form.get("especialidad"))
        tags_especialidad = strip_tags(request.form.get("tags_especialidad"))

        # Nuevos campos: Idiomas y Cursos/Certificaciones
        idiomas = strip_tags(request.form.get("idiomas"))
        cursos_certificaciones = strip_tags(request.form.get("cursos_certificaciones"))

        # Detección de scripts maliciosos en TODOS los campos
        todos_campos = [nombre, apellido_paterno, apellido_materno, ubicacion, 
                        ultimo_puesto, ultima_empresa, actividades_logros,
                        puesto_deseado, area_trabajo, especialidad, tags_especialidad,
                        idiomas, cursos_certificaciones]
        for campo in todos_campos:
            if detectar_script(campo) or detectar_sqli(campo):
                flash("⚠️ Seguridad: Se detectó contenido malicioso (Scripts/SQL) en uno de los campos. Tu solicitud fue bloqueada.")
                return redirect(url_for("registro_candidato"))

        hashed_password = generate_password_hash(password)
        conn = get_connection()
        cur = conn.cursor()
        try:
            # 1. Crear Usuario
            cur.execute("""
                INSERT INTO usuarios (id_rol, correo, password, activo) 
                VALUES (3, %s, %s, false) RETURNING id_usuario
            """, (correo, hashed_password))
            id_usuario = cur.fetchone()[0]

            # 2. Crear Candidato con todos los campos
            cur.execute("""
                INSERT INTO candidatos (
                    id_usuario, nombre, apellido_paterno, apellido_materno, 
                    sexo, telefono, codigo_postal, ubicacion,
                    carrera, ultimo_grado_estudios, institucion_estudios, carrera_estudios, anio_egreso,
                    ultimo_puesto, ultima_empresa, fecha_inicio_laboral, fecha_fin_laboral, actividades_logros,
                    puesto_deseado, sueldo_deseado, area_trabajo, especialidad, tags_especialidad,
                    idiomas, cursos_certificaciones,
                    estatus
                ) VALUES (
                    %s, %s, %s, %s,
                    %s, %s, %s, %s,
                    %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s,
                    %s, %s, %s, %s, %s,
                    %s, %s,
                    'pendiente'
                )
            """, (
                id_usuario, nombre, apellido_paterno, apellido_materno,
                sexo, telefono, codigo_postal, ubicacion,
                carrera, ultimo_grado_estudios, institucion_estudios, carrera_estudios, anio_egreso,
                ultimo_puesto, ultima_empresa, fecha_inicio_laboral, fecha_fin_laboral, actividades_logros,
                puesto_deseado, sueldo_deseado, area_trabajo, especialidad, tags_especialidad,
                idiomas, cursos_certificaciones
            ))
            
            conn.commit()
            notificar_admins("registro", f"Nuevo candidato registrado: {nombre} {apellido_paterno}", "/admin/validar-candidato")
            flash("Registro exitoso. Su cuenta será validada por un administrador.")
            return redirect(url_for("home"))
        except Exception as e:
            conn.rollback()
            print(f"Error al registrar candidato: {e}")
            flash(f"Error al registrar: {str(e)}")
            return redirect(url_for("registro_candidato"))
        finally:
            cur.close()
            conn.close()
    
    return render_template("vacantes/registro-candidato.html")
    
@app.route("/candidato-dashboard")
@login_required
def candidato_dashboard():
    if current_user.rol != "Candidato":
        flash("Acceso denegado")
        return redirect(url_for("home"))
    conn = get_connection()
    cur = conn.cursor()
    
    # Obtener id_candidato y CV
    cur.execute("SELECT id_candidato, nombre, cv_url FROM candidatos WHERE id_usuario = %s", (current_user.id,))
    candidato = cur.fetchone()
    if not candidato:
        cur.close()
        conn.close()
        flash("Perfil de candidato no encontrado.")
        return redirect(url_for("home"))
    
    id_candidato = candidato[0]
    nombre_candidato = candidato[1]
    cv_missing = candidato[2] is None
    
    # Estadísticas
    cur.execute("SELECT COUNT(*) FROM postulaciones WHERE id_candidato = %s", (id_candidato,))
    total_postulaciones = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM postulaciones WHERE id_candidato = %s AND estado = 'aceptado'", (id_candidato,))
    aceptadas = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM postulaciones WHERE id_candidato = %s AND estado = 'en revision'", (id_candidato,))
    en_revision = cur.fetchone()[0]
    
    # Actividad reciente
    cur.execute("""
        SELECT v.titulo, e.nombre, p.estado, p.fecha_postulacion
        FROM postulaciones p
        JOIN vacantes v ON p.id_vacante = v.id_vacante
        JOIN empresas e ON v.id_empresa = e.id_empresa
        WHERE p.id_candidato = %s
        ORDER BY p.fecha_postulacion DESC LIMIT 5
    """, (id_candidato,))
    actividad = cur.fetchall()
    
    cur.close()
    conn.close()
    
    stats = {
        'total_postulaciones': total_postulaciones,
        'aceptadas': aceptadas,
        'en_revision': en_revision,
        'nombre': nombre_candidato,
        'cv_missing': cv_missing
    }
    return render_template("vacantes/panel-candidato.html", stats=stats, actividad=actividad)

@app.route("/candidato/buscar")
@login_required
def candidato_buscar():
    if current_user.rol != "Candidato": return redirect(url_for("home"))
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT v.id_vacante, v.titulo, e.nombre as empresa, v.lugar_trabajo,
               v.modalidad, v.salario, v.fecha_publicacion, v.descripcion,
               v.horario, r.escolaridad, r.experiencia, r.descripcion as req_desc
        FROM vacantes v
        JOIN empresas e ON v.id_empresa = e.id_empresa
        LEFT JOIN requisitos_vacante r ON v.id_vacante = r.id_vacante
        WHERE v.estatus = 'activa' AND e.estatus = 'aprobada'
        ORDER BY v.fecha_publicacion DESC
    """)
    vacantes = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("vacantes/buscar-vacantes.html", vacantes=vacantes)

@app.route("/candidato/ver")
@login_required
def candidato_ver():
    if current_user.rol != "Candidato": return redirect(url_for("home"))
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT v.id_vacante, v.titulo, e.nombre as empresa, v.lugar_trabajo,
               v.modalidad, v.fecha_publicacion, v.estatus
        FROM vacantes v
        JOIN empresas e ON v.id_empresa = e.id_empresa
        WHERE v.estatus = 'activa' AND e.estatus = 'aprobada'
        ORDER BY v.fecha_publicacion DESC
    """)
    vacantes = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("vacantes/ver-vacantes.html", vacantes=vacantes)

@app.route("/candidato/postulaciones")
@login_required
def candidato_postulaciones():
    if current_user.rol != "Candidato": return redirect(url_for("home"))
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT id_candidato FROM candidatos WHERE id_usuario = %s", (current_user.id,))
    cand = cur.fetchone()
    if not cand:
        cur.close()
        conn.close()
        return render_template("vacantes/mis-postulaciones.html", postulaciones=[], stats={})
    
    id_candidato = cand[0]
    cur.execute("""
        SELECT v.titulo, e.nombre, p.fecha_postulacion, p.estado, p.id_postulacion, v.id_vacante, p.comentarios
        FROM postulaciones p
        JOIN vacantes v ON p.id_vacante = v.id_vacante
        JOIN empresas e ON v.id_empresa = e.id_empresa
        WHERE p.id_candidato = %s
        ORDER BY p.fecha_postulacion DESC
    """, (id_candidato,))
    postulaciones = cur.fetchall()
    
    cur.execute("SELECT COUNT(*) FROM postulaciones WHERE id_candidato = %s", (id_candidato,))
    total = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM postulaciones WHERE id_candidato = %s AND estado = 'aceptado'", (id_candidato,))
    aceptadas = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM postulaciones WHERE id_candidato = %s AND estado = 'en revision'", (id_candidato,))
    en_revision = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM postulaciones WHERE id_candidato = %s AND estado = 'rechazado'", (id_candidato,))
    rechazadas = cur.fetchone()[0]
    
    cur.close()
    conn.close()
    
    stats = {'total': total, 'aceptadas': aceptadas, 'en_revision': en_revision, 'rechazadas': rechazadas}
    return render_template("vacantes/mis-postulaciones.html", postulaciones=postulaciones, stats=stats)

@app.route("/candidato/estado-postulaciones")
@login_required
def candidato_estado_postulaciones():
    if current_user.rol != "Candidato": return redirect(url_for("home"))
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT id_candidato FROM candidatos WHERE id_usuario = %s", (current_user.id,))
    cand = cur.fetchone()
    postulaciones = []
    if cand:
        cur.execute("""
            SELECT v.titulo, e.nombre, p.fecha_postulacion, p.estado, p.comentarios
            FROM postulaciones p
            JOIN vacantes v ON p.id_vacante = v.id_vacante
            JOIN empresas e ON v.id_empresa = e.id_empresa
            WHERE p.id_candidato = %s
            ORDER BY p.fecha_postulacion DESC
        """, (cand[0],))
        postulaciones = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("vacantes/estado-postulaciones.html", postulaciones=postulaciones)

@app.route("/candidato/perfil")
@login_required
def candidato_perfil():
    if current_user.rol != "Candidato": return redirect(url_for("home"))
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT c.id_candidato, c.nombre, c.apellido_paterno, c.apellido_materno,
               c.sexo, c.telefono, c.anio_egreso, c.carrera, 
               u.correo, c.cv_url, c.codigo_postal, c.ubicacion,
               c.ultimo_grado_estudios, c.institucion_estudios, c.carrera_estudios,
               c.ultimo_puesto, c.ultima_empresa, c.fecha_inicio_laboral, c.fecha_fin_laboral,
               c.actividades_logros, c.puesto_deseado, c.sueldo_deseado,
               c.area_trabajo, c.especialidad, c.tags_especialidad, c.foto_perfil_url
        FROM candidatos c
        JOIN usuarios u ON c.id_usuario = u.id_usuario
        WHERE c.id_usuario = %s
    """, (current_user.id,))
    candidato = cur.fetchone()
    
    cur.close()
    conn.close()
    return render_template("vacantes/perfil-candidato.html", candidato=candidato)

@app.route("/candidato/subir-foto", methods=["POST"])
@login_required
def candidato_subir_foto():
    if current_user.rol != "Candidato": return redirect(url_for("home"))
    if 'foto' not in request.files:
        flash("No se seleccionó ninguna imagen.")
        return redirect(url_for("candidato_perfil"))
    
    file = request.files['foto']
    
    # Nueva Validación Fase 26
    es_valido, mensaje = validar_archivo(file, ['png', 'jpg', 'jpeg'], max_size_mb=3)
    if not es_valido:
        flash(mensaje)
        return redirect(url_for("candidato_perfil"))
    
    filename = secure_filename(f"foto_{current_user.id}_{file.filename}")
    upload_path = os.path.join("public", "uploads", "fotos", filename)
    
    # Guardar archivo
    file.save(upload_path)
    
    # Actualizar DB
    foto_url = f"/public/uploads/fotos/{filename}"
    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute("UPDATE candidatos SET foto_perfil_url = %s WHERE id_usuario = %s", (foto_url, current_user.id))
        conn.commit()
        flash("Foto de perfil actualizada exitosamente.")
    except Exception as e:
        conn.rollback()
        flash(f"Error al actualizar foto: {e}")
    finally:
        cur.close()
        conn.close()
        
    return redirect(url_for("candidato_perfil"))

@app.route("/candidato/generar-cv")
@login_required
def candidato_generar_cv():
    if current_user.rol != "Candidato": return redirect(url_for("home"))
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT c.nombre, c.apellido_paterno, c.apellido_materno,
               c.sexo, c.telefono, c.anio_egreso, c.carrera, 
               u.correo, c.codigo_postal, c.ubicacion,
               c.ultimo_grado_estudios, c.institucion_estudios, c.carrera_estudios,
               c.ultimo_puesto, c.ultima_empresa, c.fecha_inicio_laboral, c.fecha_fin_laboral,
               c.actividades_logros, c.puesto_deseado, c.sueldo_deseado,
               c.area_trabajo, c.especialidad, c.tags_especialidad, c.foto_perfil_url,
               c.id_candidato, c.idiomas, c.cursos_certificaciones
        FROM candidatos c
        JOIN usuarios u ON c.id_usuario = u.id_usuario
        WHERE c.id_usuario = %s
    """, (current_user.id,))
    candidato = cur.fetchone()
    cur.close()
    conn.close()
    return render_template("vacantes/cv-template.html", c=candidato)

@app.route("/candidato/actualizar-perfil", methods=["POST"])
@login_required
def candidato_actualizar_perfil():
    if current_user.rol != "Candidato": return redirect(url_for("home"))
    
    # Recoger todos los campos sanitizados
    nombre = strip_tags(request.form.get("nombre"))
    apellido_paterno = strip_tags(request.form.get("apellido_paterno"))
    apellido_materno = strip_tags(request.form.get("apellido_materno"))
    sexo = strip_tags(request.form.get("sexo"))
    telefono = strip_tags(request.form.get("telefono"))
    codigo_postal = strip_tags(request.form.get("codigo_postal"))
    ubicacion = strip_tags(request.form.get("ubicacion"))
    carrera = strip_tags(request.form.get("carrera"))
    egreso = request.form.get("egreso")
    egreso = int(egreso) if egreso and egreso.isdigit() else None
    ultimo_grado_estudios = strip_tags(request.form.get("ultimo_grado_estudios"))
    institucion_estudios = strip_tags(request.form.get("institucion_estudios"))
    carrera_estudios = strip_tags(request.form.get("carrera_estudios"))
    ultimo_puesto = strip_tags(request.form.get("ultimo_puesto"))
    ultima_empresa = strip_tags(request.form.get("ultima_empresa"))
    fecha_inicio_laboral = strip_tags(request.form.get("fecha_inicio_laboral"))
    fecha_fin_laboral = strip_tags(request.form.get("fecha_fin_laboral"))
    actividades_logros = strip_tags(request.form.get("actividades_logros"))
    puesto_deseado = strip_tags(request.form.get("puesto_deseado"))
    sueldo_deseado = request.form.get("sueldo_deseado")
    sueldo_deseado = float(sueldo_deseado) if sueldo_deseado and sueldo_deseado.strip() else 0
    area_trabajo = strip_tags(request.form.get("area_trabajo"))
    especialidad = strip_tags(request.form.get("especialidad"))
    tags_especialidad = strip_tags(request.form.get("tags_especialidad"))
    idiomas = strip_tags(request.form.get("idiomas"))
    cursos_certificaciones = strip_tags(request.form.get("cursos_certificaciones"))

    # Detección de scripts maliciosos
    todos_campos = [nombre, apellido_paterno, apellido_materno, ubicacion,
                    ultimo_puesto, ultima_empresa, actividades_logros,
                    puesto_deseado, area_trabajo, especialidad, tags_especialidad,
                    idiomas, cursos_certificaciones]
    for campo in todos_campos:
        if detectar_script(campo) or detectar_sqli(campo):
            flash("⚠️ Seguridad: Se detectó contenido malicioso (Scripts/SQL). Tu solicitud fue bloqueada.")
            return redirect(url_for("candidato_perfil"))

    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute("""
            UPDATE candidatos SET
                nombre = %s, apellido_paterno = %s, apellido_materno = %s,
                sexo = %s, telefono = %s, codigo_postal = %s, ubicacion = %s,
                carrera = %s, anio_egreso = %s,
                ultimo_grado_estudios = %s, institucion_estudios = %s, carrera_estudios = %s,
                ultimo_puesto = %s, ultima_empresa = %s, 
                fecha_inicio_laboral = %s, fecha_fin_laboral = %s, actividades_logros = %s,
                puesto_deseado = %s, sueldo_deseado = %s, area_trabajo = %s,
                especialidad = %s, tags_especialidad = %s,
                idiomas = %s, cursos_certificaciones = %s
            WHERE id_usuario = %s
        """, (
            nombre, apellido_paterno, apellido_materno,
            sexo, telefono, codigo_postal, ubicacion,
            carrera, egreso,
            ultimo_grado_estudios, institucion_estudios, carrera_estudios,
            ultimo_puesto, ultima_empresa,
            fecha_inicio_laboral, fecha_fin_laboral, actividades_logros,
            puesto_deseado, sueldo_deseado, area_trabajo,
            especialidad, tags_especialidad,
            idiomas, cursos_certificaciones,
            current_user.id
        ))
        conn.commit()
        flash("Perfil actualizado correctamente.")
    except Exception as e:
        conn.rollback()
        flash(f"Error al actualizar perfil: {e}")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("candidato_perfil"))

@app.route("/candidato/detalle-vacante/<int:id_vacante>")
@login_required
def candidato_detalle_vacante(id_vacante):
    if current_user.rol != "Candidato": return redirect(url_for("home"))
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT v.id_vacante, v.titulo, v.descripcion, v.salario, v.modalidad,
               v.horario, v.lugar_trabajo, v.fecha_publicacion,
               e.nombre as empresa, r.escolaridad, r.experiencia, r.descripcion as req_desc
        FROM vacantes v
        JOIN empresas e ON v.id_empresa = e.id_empresa
        LEFT JOIN requisitos_vacante r ON v.id_vacante = r.id_vacante
        WHERE v.id_vacante = %s
    """, (id_vacante,))
    vacante = cur.fetchone()
    
    # Verificar si ya se postuló
    cur.execute("SELECT id_candidato FROM candidatos WHERE id_usuario = %s", (current_user.id,))
    cand = cur.fetchone()
    ya_postulado = False
    if cand:
        cur.execute("SELECT id_postulacion FROM postulaciones WHERE id_vacante = %s AND id_candidato = %s",
                     (id_vacante, cand[0]))
        ya_postulado = cur.fetchone() is not None
    
    # Verificar si ya la guardó
    ya_guardada = False
    if cand:
        cur.execute("SELECT id_guardada FROM vacantes_guardadas WHERE id_vacante = %s AND id_candidato = %s",
                     (id_vacante, cand[0]))
        ya_guardada = cur.fetchone() is not None
        
    cur.close()
    conn.close()
    return render_template("vacantes/detalle-vacante.html", vacante=vacante, ya_postulado=ya_postulado, ya_guardada=ya_guardada)

@app.route("/candidato/postularse/<int:id_vacante>", methods=["POST"])
@login_required
def candidato_postularse(id_vacante):
    if current_user.rol != "Candidato": return redirect(url_for("home"))
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT id_candidato FROM candidatos WHERE id_usuario = %s", (current_user.id,))
    cand = cur.fetchone()
    if not cand:
        flash("Perfil de candidato no encontrado.")
        cur.close()
        conn.close()
        return redirect(url_for("candidato_dashboard"))
    
    try:
        cur.execute("""
            INSERT INTO postulaciones (id_vacante, id_candidato, estado)
            VALUES (%s, %s, 'en revision')
        """, (id_vacante, cand[0]))
        conn.commit()
        
        cur.execute("""
            SELECT u.id_usuario FROM empresas e 
            JOIN vacantes v ON e.id_empresa = v.id_empresa 
            JOIN usuarios u ON e.id_usuario = u.id_usuario
            WHERE v.id_vacante = %s
        """, (id_vacante,))
        emp_user = cur.fetchone()
        if emp_user:
            crear_notificacion(emp_user[0], "postulacion", "Nueva postulación a tu vacante.", "/empresa/consultar")
            
        notificar_admins("postulacion", f"Un candidato se ha postulado a una vacante.", "/admin/seguimiento")
        flash("Te has postulado exitosamente.")
    except Exception as e:
        conn.rollback()
        flash(f"Error al postularse: {str(e)}")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("candidato_detalle_vacante", id_vacante=id_vacante))

@app.route("/candidato/guardar-vacante/<int:id_vacante>", methods=["POST"])
@login_required
def candidato_guardar_vacante(id_vacante):
    if current_user.rol != "Candidato": return redirect(url_for("home"))
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT id_candidato FROM candidatos WHERE id_usuario = %s", (current_user.id,))
    cand = cur.fetchone()
    if not cand:
        flash("Perfil de candidato no encontrado.")
        cur.close()
        conn.close()
        return redirect(url_for("candidato_dashboard"))
        
    try:
        cur.execute("""
            INSERT INTO vacantes_guardadas (id_vacante, id_candidato)
            VALUES (%s, %s)
        """, (id_vacante, cand[0]))
        conn.commit()
        
        cur.execute("""
            SELECT u.id_usuario FROM empresas e 
            JOIN vacantes v ON e.id_empresa = v.id_empresa 
            JOIN usuarios u ON e.id_usuario = u.id_usuario
            WHERE v.id_vacante = %s
        """, (id_vacante,))
        emp_user = cur.fetchone()
        if emp_user:
            crear_notificacion(emp_user[0], "guardado", "Un candidato ha guardado tu vacante.", "/empresa")
            
        flash("Vacante guardada exitosamente.")
    except Exception as e:
        conn.rollback()
        flash(f"Error al guardar vacante: {str(e)}")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("candidato_detalle_vacante", id_vacante=id_vacante))

# ================= AUTH =================
@app.route("/auth/login", methods=["POST"])
def auth_login():
    correo = request.form.get("correo")
    password = request.form.get("password")
    rol_seleccionado = request.form.get("rol")

    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT u.id_usuario, u.correo, u.password, r.nombre AS rol, u.id_rol, u.activo
        FROM usuarios u
        JOIN roles r ON u.id_rol = r.id_rol
        WHERE u.correo = %s
    """, (correo,))
    user_data = cur.fetchone()
    cur.close()
    conn.close()

    if user_data:
        id_usuario, user_correo, hashed_password, rol_real, id_rol, is_active = user_data
        if check_password_hash(hashed_password, password):
            # Verificar si la cuenta está activa
            if not is_active and rol_real != "Administrador":
                flash("Su cuenta está pendiente de aprobación por un administrador.")
                return redirect(url_for("home"))

            if (rol_seleccionado == "admin" and rol_real == "Administrador") or \
               (rol_seleccionado == "empresa" and rol_real == "Empresa") or \
               (rol_seleccionado == "candidato" and rol_real == "Candidato"):
                # === 2FA: Generar y enviar código ===
                code_2fa = str(random.randint(100000, 999999))
                expires_2fa = datetime.now() + timedelta(minutes=5)
                
                conn = get_connection()
                cur = conn.cursor()
                cur.execute("""
                    UPDATE usuarios 
                    SET two_fa_code = %s, two_fa_expires = %s 
                    WHERE id_usuario = %s
                """, (code_2fa, expires_2fa, id_usuario))
                conn.commit()
                cur.close()
                conn.close()

                session['2fa_user_id'] = id_usuario
                session['2fa_correo'] = user_correo
                session['2fa_rol'] = rol_real
                session['2fa_id_rol'] = id_rol
                
                # Enviar código por email
                asunto = "Código de Verificación - UT Oriental Emplea"
                mensaje = (f"Hola,\n\nTu código de verificación es: {code_2fa}\n\n"
                          f"Este código expira en 5 minutos.\n\n"
                          f"Si no solicitaste este código, ignora este mensaje.\n\n"
                          f"Saludos,\nUT Oriental Emplea")
                enviar_email(user_correo, asunto, mensaje)
                
                flash("Se ha enviado un código de verificación a tu correo electrónico.")
                return redirect(url_for("auth_verify_2fa"))
            else:
                flash("El rol seleccionado no coincide con su cuenta.")
                return redirect(url_for("home"))
        else:
            flash("Contraseña incorrecta.")
            return redirect(url_for("home"))
    else:
        flash("El correo no está registrado.")
        return redirect(url_for("home"))

# ================= 2FA VERIFICATION =================
@app.route("/auth/verify-2fa", methods=["GET"])
def auth_verify_2fa():
    if '2fa_user_id' not in session:
        flash("Sesión de verificación expirada. Inicia sesión de nuevo.")
        return redirect(url_for("home"))
    correo_parcial = session.get('2fa_correo', '')
    if '@' in correo_parcial:
        parts = correo_parcial.split('@')
        correo_parcial = parts[0][:3] + '***@' + parts[1]
    return render_template("home/verify_2fa.html", correo_parcial=correo_parcial)

@app.route("/auth/verify-2fa", methods=["POST"])
def auth_verify_2fa_post():
    codigo_ingresado = request.form.get("codigo", "").strip()
    
    # Verificar en base de datos
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT two_fa_code, two_fa_expires FROM usuarios WHERE id_usuario = %s", (session['2fa_user_id'],))
    db_2fa = cur.fetchone()
    cur.close()
    conn.close()

    if not db_2fa or not db_2fa[0]:
        flash("No se encontró un código de verificación activo.")
        return redirect(url_for("home"))

    db_code, db_expires = db_2fa
    
    # Verificar expiración
    if datetime.now() > db_expires:
        flash("El código ha expirado. Inicia sesión de nuevo.")
        return redirect(url_for("home"))
    
    if codigo_ingresado == db_code:
        # Código correcto → completar login
        id_usuario = session['2fa_user_id']
        correo = session['2fa_correo']
        id_rol = session['2fa_id_rol']
        rol_real = session['2fa_rol']
        
        # Limpiar datos 2FA de la sesión y BD
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("UPDATE usuarios SET two_fa_code = NULL, two_fa_expires = NULL WHERE id_usuario = %s", (id_usuario,))
        conn.commit()
        cur.close()
        conn.close()

        session.pop('2fa_user_id', None)
        session.pop('2fa_correo', None)
        session.pop('2fa_rol', None)
        session.pop('2fa_id_rol', None)
        
        user_obj = User(id_usuario, correo, id_rol, rol_real)
        login_user(user_obj)
        
        if rol_real == "Administrador":
            return redirect(url_for("admin"))
        elif rol_real == "Empresa":
            return redirect(url_for("empresa.inicio"))
        else:
            return redirect(url_for("candidato_dashboard"))
    else:
        flash("Código incorrecto. Inténtalo de nuevo.")
        return redirect(url_for("auth_verify_2fa"))

@app.route("/auth/resend-2fa", methods=["POST"])
def auth_resend_2fa():
    if '2fa_correo' not in session:
        flash("Sesión expirada. Inicia sesión de nuevo.")
        return redirect(url_for("home"))
    
    code_2fa = str(random.randint(100000, 999999))
    expires_2fa = datetime.now() + timedelta(minutes=5)
    
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("UPDATE usuarios SET two_fa_code = %s, two_fa_expires = %s WHERE id_usuario = %s", 
                (code_2fa, expires_2fa, session['2fa_user_id']))
    conn.commit()
    cur.close()
    conn.close()
    
    correo = session['2fa_correo']
    asunto = "Nuevo Código de Verificación - UT Oriental Emplea"
    mensaje = (f"Hola,\n\nTu nuevo código de verificación es: {code_2fa}\n\n"
              f"Este código expira en 5 minutos.\n\n"
              f"Saludos,\nUT Oriental Emplea")
    enviar_email(correo, asunto, mensaje)
    
    flash("Se ha enviado un nuevo código a tu correo.")
    return redirect(url_for("auth_verify_2fa"))

# ================= FORGOT PASSWORD =================
@app.route("/auth/forgot-password", methods=["POST"])
def auth_forgot_password():
    correo = request.form.get("correo", "").strip()
    if not correo:
        flash("Por favor ingresa tu correo electrónico.")
        return redirect(url_for("home"))
    
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT id_usuario FROM usuarios WHERE correo = %s", (correo,))
    user = cur.fetchone()
    cur.close()
    conn.close()
    
    if user:
        # Generar token y almacenar en DB
        token = str(uuid.uuid4())
        expires = datetime.now() + timedelta(hours=1)
        
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            UPDATE usuarios SET reset_token = %s, reset_expires = %s 
            WHERE id_usuario = %s
        """, (token, expires, user[0]))
        conn.commit()
        cur.close()
        conn.close()
        
        # Enviar email con link de reset
        reset_url = url_for('auth_reset_password', token=token, _external=True)
        asunto = "Restablecer Contraseña - UT Oriental Emplea"
        mensaje = (f"Hola,\n\nRecibimos una solicitud para restablecer tu contraseña.\n\n"
                  f"Haz clic en el siguiente enlace para crear una nueva contraseña:\n\n"
                  f"{reset_url}\n\n"
                  f"Este enlace expira en 1 hora.\n\n"
                  f"Si no solicitaste este cambio, ignora este mensaje.\n\n"
                  f"Saludos,\nUT Oriental Emplea")
        enviar_email(correo, asunto, mensaje)
    
    # Siempre mostrar mismo mensaje por seguridad
    flash("Si el correo está registrado, recibirás un enlace para restablecer tu contraseña.")
    return redirect(url_for("home"))

@app.route("/auth/reset-password/<token>", methods=["GET"])
def auth_reset_password(token):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT id_usuario, reset_expires FROM usuarios WHERE reset_token = %s", (token,))
    user = cur.fetchone()
    cur.close()
    conn.close()

    if not user:
        flash("El enlace de recuperación no es válido o ha expirado.")
        return redirect(url_for("home"))
    
    expires = user[1]
    if datetime.now() > expires:
        flash("El enlace de recuperación ha expirado. Solicita uno nuevo.")
        return redirect(url_for("home"))
    
    return render_template("home/reset_password.html", token=token)

@app.route("/auth/reset-password/<token>", methods=["POST"])
def auth_reset_password_post(token):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT id_usuario, reset_expires FROM usuarios WHERE reset_token = %s", (token,))
    user = cur.fetchone()
    
    if not user:
        cur.close()
        conn.close()
        flash("El enlace de recuperación no es válido o ha expirado.")
        return redirect(url_for("home"))
    
    user_id, expires = user
    if datetime.now() > expires:
        cur.close()
        conn.close()
        flash("El enlace de recuperación ha expirado.")
        return redirect(url_for("home"))
    
    new_password = request.form.get("password", "")
    confirm_password = request.form.get("confirm_password", "")
    
    if new_password != confirm_password:
        flash("Las contraseñas no coinciden.")
        return redirect(url_for("auth_reset_password", token=token))
    
    # Validar contraseña
    if len(new_password) < 12:
        flash("La contraseña debe tener al menos 12 caracteres.")
        return redirect(url_for("auth_reset_password", token=token))
    if not re.search(r"[A-Z]", new_password) or not re.search(r"[a-z]", new_password) or \
       not re.search(r"\d", new_password) or not re.search(r"[@$!%*?&]", new_password):
        flash("La contraseña debe incluir mayúscula, minúscula, número y símbolo especial.")
        return redirect(url_for("auth_reset_password", token=token))
    
    hashed = generate_password_hash(new_password)
    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute("""
            UPDATE usuarios 
            SET password = %s, reset_token = NULL, reset_expires = NULL 
            WHERE id_usuario = %s
        """, (hashed, user_id))
        conn.commit()
        flash("Contraseña actualizada exitosamente. Ya puedes iniciar sesión.")
    except Exception as e:
        conn.rollback()
        flash(f"Error al actualizar la contraseña: {e}")
    finally:
        cur.close()
        conn.close()
    
    return redirect(url_for("home"))

# ================= API ENDPOINTS =================
@app.route("/api/vacante/<int:id_vacante>")
@login_required
def api_get_vacante(id_vacante):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT v.id_vacante, v.titulo, v.descripcion, v.salario, v.modalidad,
               v.horario, v.lugar_trabajo, v.fecha_publicacion,
               e.nombre as empresa, r.escolaridad, r.experiencia, r.descripcion as req_desc
        FROM vacantes v
        JOIN empresas e ON v.id_empresa = e.id_empresa
        LEFT JOIN requisitos_vacante r ON v.id_vacante = r.id_vacante
        WHERE v.id_vacante = %s
    """, (id_vacante,))
    row = cur.fetchone()
    cur.close()
    conn.close()
    
    if not row:
        return {"error": "Vacante no encontrada"}, 404
        
    return {
        "id_vacante": row[0],
        "titulo": row[1],
        "descripcion": row[2],
        "salario": f"${row[3]:,.2f}" if row[3] else "No especificado",
        "modalidad": row[4],
        "horario": row[5],
        "lugar": row[6],
        "fecha": row[7].strftime('%d/%m/%y') if row[7] else "N/A",
        "empresa": row[8],
        "escolaridad": row[9] or "No especificada",
        "experiencia": row[10] or "No especificada",
        "requisitos": row[11] or "No especificados"
    }

@app.route("/logout")
@login_required
def logout():
    logout_user()
    flash("Sesión cerrada correctamente.")
    return redirect(url_for("home"))

# ================= CV UPLOAD / DOWNLOAD =================
@app.route("/candidato/subir-cv", methods=["POST"])
@login_required
def candidato_subir_cv():
    if current_user.rol != "Candidato": return redirect(url_for("home"))
    if 'cv' not in request.files:
        flash("No se seleccionó ningún archivo.")
        return redirect(url_for("candidato_perfil"))
    
    file = request.files['cv']
    
    # Nueva Validación Fase 26
    es_valido, mensaje = validar_archivo(file, ['pdf', 'doc', 'docx'], max_size_mb=5)
    if not es_valido:
        flash(mensaje)
        return redirect(url_for("candidato_perfil"))
    
    filename = secure_filename(f"cv_{current_user.id}_{file.filename}")
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(filepath)
    
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("UPDATE candidatos SET cv_url = %s WHERE id_usuario = %s", (filename, current_user.id))
    conn.commit()
    crear_notificacion(current_user.id, "sistema", "Has actualizado tu CV exitosamente.", "/candidato/perfil")
    cur.close()
    conn.close()
    flash("CV subido exitosamente.")
    return redirect(url_for("candidato_perfil"))

@app.route("/ver-cv/<int:id_candidato>")
@login_required
def ver_cv(id_candidato):
    conn = get_connection()
    cur = conn.cursor()
    
    # 1. Obtener datos del candidato y su id_usuario (dueño)
    cur.execute("SELECT cv_url, id_usuario FROM candidatos WHERE id_candidato = %s", (id_candidato,))
    result = cur.fetchone()
    
    if not result or not result[0]:
        cur.close()
        conn.close()
        flash("El candidato no ha subido su CV aún.")
        return redirect(request.referrer or url_for("home"))
    
    cv_url, owner_id = result
    authorized = False
    
    # 2. Autorización por Rol
    if current_user.rol == 'Administrador':
        authorized = True
    elif current_user.id == owner_id:
        authorized = True
    elif current_user.rol == 'Empresa':
        # Verificar si hay una postulación del candidato a esta empresa
        cur.execute("""
            SELECT 1 FROM postulaciones p
            JOIN vacantes v ON p.id_vacante = v.id_vacante
            JOIN empresas e ON v.id_empresa = e.id_empresa
            WHERE p.id_candidato = %s AND e.id_usuario = %s
            LIMIT 1
        """, (id_candidato, current_user.id))
        if cur.fetchone():
            authorized = True
            crear_notificacion(owner_id, "vista", "Una empresa ha visto tu perfil/CV.", "/candidato/estado-postulaciones")

    cur.close()
    conn.close()
    
    if authorized:
        return send_from_directory(app.config['UPLOAD_FOLDER'], cv_url, as_attachment=False)
    
    flash("No tienes permiso para ver este currículum.")
    return redirect(url_for("home"))

@app.route("/descargar-cv/<int:id_candidato>")
@login_required
def descargar_cv(id_candidato):
    # Reutilizamos la lógica de ver_cv o implementamos similar
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT cv_url, id_usuario FROM candidatos WHERE id_candidato = %s", (id_candidato,))
    result = cur.fetchone()
    
    if not result or not result[0]:
        cur.close()
        conn.close()
        flash("El candidato no ha subido su CV aún.")
        return redirect(request.referrer or url_for("home"))
    
    cv_url, owner_id = result
    authorized = False
    
    if current_user.rol == 'Administrador' or current_user.id == owner_id:
        authorized = True
    elif current_user.rol == 'Empresa':
        cur.execute("""
            SELECT 1 FROM postulaciones p
            JOIN vacantes v ON p.id_vacante = v.id_vacante
            JOIN empresas e ON v.id_empresa = e.id_empresa
            WHERE p.id_candidato = %s AND e.id_usuario = %s
            LIMIT 1
        """, (id_candidato, current_user.id))
        if cur.fetchone():
            authorized = True

    cur.close()
    conn.close()
    
    if authorized:
        return send_from_directory(app.config['UPLOAD_FOLDER'], cv_url, as_attachment=True)
    
    flash("No tienes permiso para descargar este currículum.")
    return redirect(url_for("home"))

# ================= ADMIN: DETALLE CANDIDATO =================
@app.route("/admin/detalle-candidato/<int:id_candidato>")
@login_required
def admin_detalle_candidato(id_candidato):
    if current_user.rol != "Administrador": return redirect(url_for("home"))
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        SELECT c.id_candidato, c.nombre, c.apellido_paterno, c.apellido_materno,
               c.sexo, c.telefono, c.tipo_usuario, c.anio_egreso, c.cv_url,
               c.fecha_registro, c.estatus, u.correo,
               c.carrera
        FROM candidatos c
        JOIN usuarios u ON c.id_usuario = u.id_usuario
        WHERE c.id_candidato = %s
    """, (id_candidato,))
    candidato = cur.fetchone()
    
    # Postulaciones del candidato
    cur.execute("""
        SELECT v.titulo, e.nombre, p.fecha_postulacion, p.estado
        FROM postulaciones p
        JOIN vacantes v ON p.id_vacante = v.id_vacante
        JOIN empresas e ON v.id_empresa = e.id_empresa
        WHERE p.id_candidato = %s
        ORDER BY p.fecha_postulacion DESC
    """, (id_candidato,))
    postulaciones = cur.fetchall()
    
    cur.close()
    conn.close()
    return render_template("admi/detalle_candidato.html", candidato=candidato, postulaciones=postulaciones)

# ================= EMPRESA: GESTIÓN DE VACANTES =================

@app.route("/empresa/vacante/<int:id_vacante>/ver")
@login_required
def empresa_ver_vacante(id_vacante):
    if current_user.rol != "Empresa": return redirect(url_for("home"))
    conn = get_connection()
    cur = conn.cursor()
    
    # Obtener detalles completos de la vacante
    cur.execute("""
        SELECT v.id_vacante, v.titulo, v.descripcion, v.salario, v.modalidad, 
               v.horario, v.lugar_trabajo, v.fecha_vencimiento, v.num_vacantes, v.perfil,
               r.escolaridad, r.experiencia, r.descripcion as req_desc
        FROM vacantes v
        LEFT JOIN requisitos_vacante r ON v.id_vacante = r.id_vacante
        JOIN empresas e ON v.id_empresa = e.id_empresa
        WHERE v.id_vacante = %s AND e.id_usuario = %s
    """, (id_vacante, current_user.id))
    vacante = cur.fetchone()
    
    if not vacante:
        cur.close()
        conn.close()
        flash("Vacante no encontrada o sin acceso.")
        return redirect(url_for("empresa.consultar"))
    
    # Obtener postulaciones
    cur.execute("""
        SELECT p.id_postulacion, c.id_candidato, c.nombre, u.correo, c.telefono,
               p.fecha_postulacion, p.estado, c.cv_url,
               COALESCE(ca.nombre, 'N/A') as carrera, u.id_usuario
        FROM postulaciones p
        JOIN candidatos c ON p.id_candidato = c.id_candidato
        JOIN usuarios u ON c.id_usuario = u.id_usuario
        LEFT JOIN carreras ca ON c.id_carrera = ca.id_carrera
        WHERE p.id_vacante = %s
        ORDER BY p.fecha_postulacion DESC
    """, (id_vacante,))
    postulaciones_db = cur.fetchall()
    
    # Notify those who are "en revision" and change to "visto"
    postulaciones = []
    for p in postulaciones_db:
        estado_actual = p[6]
        if estado_actual == 'en revision':
            cur.execute("UPDATE postulaciones SET estado = 'visto' WHERE id_postulacion = %s", (p[0],))
            crear_notificacion(p[9], "vista", f"La empresa está revisando tu postulación a {vacante[1]}.", "/candidato/estado-postulaciones")
            estado_actual = 'visto'
        
        # Build the tuple back (tuples are immutable so we construct a dict or fake it, but template uses indices)
        # Template uses indices 0..8 usually. Let's make a new list of elements
        new_p = list(p)
        new_p[6] = estado_actual
        postulaciones.append(new_p)
    
    conn.commit()
    cur.close()
    conn.close()
    return render_template("empresa/empresa-postulaciones.html", vacante=vacante, postulaciones=postulaciones)

@app.route("/empresa/vacante/<int:id_vacante>/editar", methods=["GET", "POST"])
@login_required
def empresa_editar_vacante(id_vacante):
    if current_user.rol != "Empresa": return redirect(url_for("home"))
    conn = get_connection()
    cur = conn.cursor()
    
    if request.method == "POST":
        titulo = request.form.get("titulo")
        descripcion = request.form.get("descripcion")
        salario = request.form.get("salario")
        modalidad = request.form.get("modalidad")
        horario = request.form.get("horario")
        lugar = request.form.get("lugar")
        fecha_vencimiento = request.form.get("fecha_vencimiento")
        escolaridad = request.form.get("escolaridad")
        experiencia = request.form.get("experiencia")
        conocimientos = request.form.get("conocimientos")
        num_vacantes = request.form.get("num_vacantes")
        perfil = request.form.get("perfil")
        
        try:
            cur.execute("""
                UPDATE vacantes SET titulo=%s, descripcion=%s, salario=%s, modalidad=%s, 
                horario=%s, lugar_trabajo=%s, fecha_vencimiento=%s, num_vacantes=%s, perfil=%s
                WHERE id_vacante = %s
            """, (titulo, descripcion, float(salario) if salario else None, modalidad, horario, lugar, fecha_vencimiento, int(num_vacantes) if num_vacantes else 1, perfil, id_vacante))
            
            cur.execute("""
                UPDATE requisitos_vacante SET escolaridad=%s, experiencia=%s, descripcion=%s
                WHERE id_vacante = %s
            """, (escolaridad, experiencia, conocimientos, id_vacante))
            
            conn.commit()
            flash("Vacante actualizada exitosamente.")
            return redirect(url_for("empresa.consultar"))
        except Exception as e:
            conn.rollback()
            flash(f"Error al actualizar: {str(e)}")
    
    # GET: Cargar datos actuales
    cur.execute("""
        SELECT v.id_vacante, v.titulo, v.descripcion, v.salario, v.modalidad, 
               v.horario, v.lugar_trabajo, v.fecha_vencimiento, v.num_vacantes, v.perfil,
               r.escolaridad, r.experiencia, r.descripcion as req_desc
        FROM vacantes v
        LEFT JOIN requisitos_vacante r ON v.id_vacante = r.id_vacante
        WHERE v.id_vacante = %s
    """, (id_vacante,))
    vacante = cur.fetchone()
    cur.close()
    conn.close()
    
    return render_template("empresa/empresa-editar-vacante.html", vacante=vacante)

@app.route("/empresa/vacante/<int:id_vacante>/eliminar", methods=["POST"])
@login_required
def empresa_eliminar_vacante(id_vacante):
    if current_user.rol != "Empresa": return redirect(url_for("home"))
    conn = get_connection()
    cur = conn.cursor()
    try:
        # Primero eliminar requisitos (por FK)
        cur.execute("DELETE FROM requisitos_vacante WHERE id_vacante = %s", (id_vacante,))
        # Eliminar postulaciones (opcional o obligatorio dependiendo de la lógica, 
        # usualmente si hay postulaciones no se debería eliminar tan fácil, 
        # pero para esta tarea lo haremos simple)
        cur.execute("DELETE FROM postulaciones WHERE id_vacante = %s", (id_vacante,))
        cur.execute("DELETE FROM vacantes WHERE id_vacante = %s", (id_vacante,))
        conn.commit()
        notificar_admins("vacantes", "Una empresa ha eliminado una vacante.", "/admin/vacantes")
        flash("Vacante eliminada.")
    except Exception as e:
        conn.rollback()
        flash(f"Error al eliminar: {str(e)}")
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("empresa.consultar"))

@app.route("/empresa/postulacion/<int:id_postulacion>/aceptar", methods=["POST"])
@login_required
def empresa_aceptar_postulacion(id_postulacion):
    if current_user.rol != "Empresa": return redirect(url_for("home"))
    comentario = request.form.get("comentario", "")
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("UPDATE postulaciones SET estado = 'aceptado', comentarios = %s WHERE id_postulacion = %s", (comentario, id_postulacion))
    cur.execute("SELECT id_vacante, id_candidato FROM postulaciones WHERE id_postulacion = %s", (id_postulacion,))
    result = cur.fetchone()
    vacante = (result[0],) if result else None
    if result:
        cur.execute("SELECT id_usuario FROM candidatos WHERE id_candidato = %s", (result[1],))
        cand_user = cur.fetchone()
        if cand_user:
            msg = "Has sido aceptado en una vacante."
            if comentario: msg += f" Mensaje: {comentario}"
            
            asunto_c = "¡Felicidades! Has sido aceptado - UT Oriental Emplea"
            cuerpo_c = f"Hola,\n\nBuenas noticias: Una empresa ha aceptado tu postulación.\n\n{msg}\n\nIngresa al portal para ver más detalles."
            crear_notificacion(cand_user[0], "aceptado", msg, "/candidato/estado-postulaciones", asunto_c, cuerpo_c)
        notificar_admins("seleccion", "Una empresa ha aceptado a un candidato.", "/admin/seguimiento")
    conn.commit()
    cur.close()
    conn.close()
    flash("Candidato aceptado con éxito.")
    if vacante:
        return redirect(url_for("empresa_ver_vacante", id_vacante=vacante[0]))
    return redirect(url_for("empresa.consultar"))

@app.route("/empresa/postulacion/<int:id_postulacion>/rechazar", methods=["POST"])
@login_required
def empresa_rechazar_postulacion(id_postulacion):
    if current_user.rol != "Empresa": return redirect(url_for("home"))
    comentario = request.form.get("comentario", "")
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("UPDATE postulaciones SET estado = 'rechazado', comentarios = %s WHERE id_postulacion = %s", (comentario, id_postulacion))
    cur.execute("SELECT id_vacante, id_candidato FROM postulaciones WHERE id_postulacion = %s", (id_postulacion,))
    result = cur.fetchone()
    vacante = (result[0],) if result else None
    if result:
        cur.execute("SELECT id_usuario FROM candidatos WHERE id_candidato = %s", (result[1],))
        cand_user = cur.fetchone()
        if cand_user:
            msg = "Tu postulación ha sido rechazada."
            if comentario: msg += f" Motivo: {comentario}"
            
            asunto_r = "Actualización de Postulación - UT Oriental Emplea"
            cuerpo_r = f"Hola,\n\nSe ha actualizado el estado de tu postulación.\n\n{msg}\n\nTe deseamos éxito en otras vacantes."
            crear_notificacion(cand_user[0], "rechazado", msg, "/candidato/estado-postulaciones", asunto_r, cuerpo_r)
        notificar_admins("seleccion", "Una empresa ha rechazado a un candidato.", "/admin/seguimiento")
    conn.commit()
    cur.close()
    conn.close()
    flash("Candidato rechazado con éxito.")
    if vacante:
        return redirect(url_for("empresa_ver_vacante", id_vacante=vacante[0]))
    return redirect(url_for("empresa.consultar"))

# ================= MAIN =================
if __name__ == "__main__":
    app.run(host='0.0.0.0', port=5001, debug=True)
